#!/usr/bin/env python3
"""Build the comprehensive triggers + sentences review Excel for Roman.

Output: scanner/explore_v2/scanner_copy_review_2026_04_27.xlsx

One sheet per section. Each row: ID, Trigger, Sentence template, Concrete
example, Mentions shoe model?, Origin (V1/V2/Proposed), Status, Notes,
ROMAN_COMMENT (empty column for review).

Sandbox-only: writes one .xlsx file. No production / DB touches.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from pathlib import Path

OUT = Path(__file__).resolve().parent / "scanner_copy_review_2026_04_27.xlsx"

# ─── Style constants ────────────────────────────────────────────
FONT_NAME = "Arial"

# Status colors (for the Status column cell + row tint)
STATUS_FILL = {
    "LOCKED":   PatternFill("solid", start_color="C6EFCE"),  # green
    "PROPOSED": PatternFill("solid", start_color="FFF2CC"),  # yellow
    "V1-LIVE":  PatternFill("solid", start_color="DDEBF7"),  # blue
    "V2-DRAFT": PatternFill("solid", start_color="FCE4D6"),  # orange
    "DROPPED":  PatternFill("solid", start_color="F2F2F2"),  # grey
}

HEADER_FILL = PatternFill("solid", start_color="2C3227")  # dark earth tone
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
COVER_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color="2C3227")
COVER_BODY_FONT = Font(name=FONT_NAME, size=11, color="2C3227")
THIN_BORDER = Border(
    left=Side(style="thin", color="D5CDBF"),
    right=Side(style="thin", color="D5CDBF"),
    top=Side(style="thin", color="D5CDBF"),
    bottom=Side(style="thin", color="D5CDBF"),
)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
TOP_LEFT = Alignment(vertical="top", horizontal="left")
TOP_CENTER = Alignment(vertical="top", horizontal="center")

COLS = [
    ("ID",                10),
    ("Trigger",           48),
    ("Sentence template", 60),
    ("Concrete example",  60),
    ("Mentions shoe?",    11),
    ("Origin",            12),
    ("Status",            13),
    ("Notes",             40),
    ("ROMAN COMMENT",     35),
]

# ─── Helpers ────────────────────────────────────────────────────
def write_header(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color="2C3227")
    ws.merge_cells("A1:I1")
    ws["A1"].alignment = TOP_LEFT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name=FONT_NAME, italic=True, size=10, color="7A7462")
        ws.merge_cells("A2:I2")
        ws["A2"].alignment = TOP_LEFT
    # Column header row at row 4
    for ci, (label, width) in enumerate(COLS, start=1):
        c = ws.cell(row=4, column=ci, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = "A5"


def add_row(ws, row_num, vals):
    """vals: dict with keys matching COLS labels. status drives row fill."""
    status = vals.get("Status", "")
    fill = STATUS_FILL.get(status)
    for ci, (label, _) in enumerate(COLS, start=1):
        v = vals.get(label, "")
        c = ws.cell(row=row_num, column=ci, value=v)
        c.font = Font(name=FONT_NAME, size=10, color="2C3227")
        c.alignment = WRAP
        c.border = THIN_BORDER
        if label == "Status" and fill:
            c.fill = fill
            c.font = Font(name=FONT_NAME, size=10, bold=True, color="2C3227")
            c.alignment = TOP_CENTER
        elif label == "Mentions shoe?" and v in ("Yes", "No"):
            c.alignment = TOP_CENTER
        elif label == "Origin":
            c.alignment = TOP_CENTER


def auto_row_height(ws, start_row, end_row):
    """Rough heuristic: longer text => taller row."""
    for r in range(start_row, end_row + 1):
        max_len = 0
        for col_idx in (3, 4):  # Sentence template + Concrete example
            v = ws.cell(row=r, column=col_idx).value or ""
            max_len = max(max_len, len(str(v)))
        # ~60 chars/line at column width ~60 => need ceil(max_len/60) lines
        lines = max(1, (max_len // 55) + 1)
        ws.row_dimensions[r].height = max(20, 14 * lines)


# ─── Cover sheet ────────────────────────────────────────────────
def build_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = "Scanner copy review — triggers + sentences"
    ws["A1"].font = COVER_TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = "Generated 2026-04-27 for in-flight review (Roman). Sandbox-only."
    ws["A2"].font = Font(name=FONT_NAME, italic=True, size=11, color="7A7462")
    ws.merge_cells("A2:H2")

    ws["A4"] = "How to read this workbook"
    ws["A4"].font = Font(name=FONT_NAME, bold=True, size=13, color="2C3227")
    intro = [
        "Each sheet covers one section of the scanner output. Rows = individual paragraphs / sentences the engine can emit.",
        "Use the 'ROMAN COMMENT' column on the right to mark up edits, deletions, or new triggers you want.",
        "",
        "Status column legend:",
        "  • LOCKED   — design already agreed, implemented in the V2 sandbox.",
        "  • PROPOSED — my draft, waiting on your sign-off before implementation.",
        "  • V1-LIVE  — what production currently emits (live at /scan today).",
        "  • V2-DRAFT — what V2 sandbox currently emits, may differ from V1.",
        "  • DROPPED  — V1 paragraph that V2 removes (kept here so you can flag if it should stay).",
        "",
        "Sheet order:",
        "  1. §1 Your Foot Shape — LOCKED design (T2 + T4), implemented today.",
        "  2. §2 Shoe Fit Tells Us — S1 locked, S2/S3/S4/S5 proposed for review.",
        "  3. §3 What to Look For — full V1+V2 inventory + cleanup proposals.",
        "  4. Shoe Card P1 (description) — sentences on each recommendation card.",
        "  5. Shoe Card P2 (why selected) — selection rationale per shoe.",
        "  6. Shoe Card P3 (tradeoffs) — flagged compromises per shoe.",
        "  7. Cross-cutting — sizing intro, tier headers, anchor labels.",
        "",
        "Where the sentences live in code:",
        "  scanner/benchmark/interp_foot_shape.py            -- V1 §1",
        "  scanner/benchmark/interp_shoe_fit.py              -- V1 §2",
        "  scanner/benchmark/interp_what_to_look_for.py      -- V1 §3",
        "  scanner/benchmark/interp_shoe_desc.py             -- V1 P1/P2/P3",
        "  scanner/explore_v2/interp_foot_shape_v2.py        -- V2 §1 (sandbox)",
        "  scanner/explore_v2/interp_shoe_fit_v2.py          -- V2 §2 (sandbox)",
        "  scanner/explore_v2/interp_what_to_look_for_v2.py  -- V2 §3 (sandbox)",
        "  scanner/explore_v2/interp_shoe_desc_v2.py         -- V2 P1/P2/P3 (sandbox)",
    ]
    for i, line in enumerate(intro, start=5):
        c = ws.cell(row=i, column=1, value=line)
        c.font = COVER_BODY_FONT
        c.alignment = TOP_LEFT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    legend_header_row = 5 + len(intro) + 1
    legend_row_start = legend_header_row + 1

    # Status legend coloured boxes
    hdr = ws.cell(row=legend_header_row, column=1, value="Status colour key:")
    hdr.font = Font(name=FONT_NAME, bold=True, size=12, color="2C3227")
    statuses = [
        ("LOCKED", "Implemented in V2 sandbox; final wording agreed."),
        ("PROPOSED", "New design, awaiting your sign-off."),
        ("V1-LIVE", "Currently emitted by production /scan; review for V2 retention."),
        ("V2-DRAFT", "Currently emitted by V2 sandbox; review for cutover."),
        ("DROPPED", "V1 paragraph removed from V2; flag if it should come back."),
    ]
    for i, (status, desc) in enumerate(statuses, start=legend_row_start):
        sc = ws.cell(row=i, column=1, value=status)
        sc.font = Font(name=FONT_NAME, bold=True, size=11, color="2C3227")
        sc.fill = STATUS_FILL[status]
        sc.alignment = TOP_CENTER
        sc.border = THIN_BORDER
        dc = ws.cell(row=i, column=2, value=desc)
        dc.font = COVER_BODY_FONT
        dc.alignment = TOP_LEFT
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)

    # Column widths
    ws.column_dimensions["A"].width = 14
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16
    ws.row_dimensions[1].height = 28


# ─── §1 sheet ───────────────────────────────────────────────────
def build_section1(wb):
    ws = wb.create_sheet("§1 Foot Shape")
    write_header(ws, "Section 1 — Your Foot Shape",
                 "LOCKED design as of 2026-04-27. Implemented in interp_foot_shape_v2.py. 2 paragraphs total.")

    rows = [
        # T2 — Basics paragraph
        dict(ID="T2.1", Trigger="Always (any scan with toe shape + width data)",
             **{
                "Sentence template": 'You have {toe_article} {Toe} toe form and {forefoot_width} forefoot with a {heel_width} heel.',
                "Concrete example": 'You have an Egyptian toe form and narrow forefoot with a narrow heel.',
                "Mentions shoe?": "No",
                "Origin": "V2 NEW",
                "Status": "LOCKED",
                "Notes": 'toe_article: "an" for Egyptian, "a" for Greek/Roman. forefoot_width / heel_width: narrow / normal / wide (soft-classified, ±0.005 tolerance to nearest tertile).',
                "ROMAN COMMENT": "",
            }),
        dict(ID="T3.1", Trigger="forefoot_width == narrow AND heel_width == narrow heel",
             **{
                "Sentence template": "Throughout a narrow profile.",
                "Concrete example": "Throughout a narrow profile.",
                "Mentions shoe?": "No",
                "Origin": "V2 NEW",
                "Status": "LOCKED",
                "Notes": "Sentence 2 of T2. T3.* picks the overall profile label.",
                "ROMAN COMMENT": "",
            }),
        dict(ID="T3.2", Trigger="forefoot_width == wide AND heel_width == wide heel",
             **{"Sentence template": "Throughout a wide profile.",
                "Concrete example": "Throughout a wide profile.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T3.3", Trigger="forefoot_width == normal AND heel_width == normal",
             **{"Sentence template": "Throughout a medium profile.",
                "Concrete example": "Throughout a medium profile.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T3.4", Trigger="Any other forefoot×heel combo (narrow+normal / normal+narrow / wide+normal / normal+wide / narrow+wide / wide+narrow)",
             **{"Sentence template": "A mixed profile; forefoot and heel require different fits.",
                "Concrete example": "A mixed profile; forefoot and heel require different fits.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "Collapses 6 combinations into one sentence.", "ROMAN COMMENT": ""}),

        # T4 — Secondaries paragraph
        dict(ID="T4.0", Trigger="META rule (always emits one paragraph)",
             **{"Sentence template": "If NONE of T4.2 .. T4.10 trigger -> emit T4.1. Otherwise emit triggered T4.x sentences in order T4.2 -> T4.10, separated by spaces, prefixed with 'Beyond the obvious: '.",
                "Concrete example": "(structural rule, not a sentence)",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "Single rule. No 'all-average sentence trigger OR something else' — clean if/else.", "ROMAN COMMENT": ""}),
        dict(ID="T4.1", Trigger="toe == Egyptian AND arch == normal AND instep == normal AND heel_depth == normal AND HVA == normal",
             **{"Sentence template": "As your arch length, instep height and heel depth are within average, you can focus mostly on toe form and shoe width when looking for a new shoe.",
                "Concrete example": "As your arch length, instep height and heel depth are within average, you can focus mostly on toe form and shoe width when looking for a new shoe.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "The 'all average' fallback when nothing notable.", "ROMAN COMMENT": ""}),
        dict(ID="T4.2", Trigger="toe IN (Greek, Roman)",
             **{"Sentence template": "Given your {toe_lower} toes, rather avoid very pointy shoes. These might squeeze your toe tips, even if the width of the shoe fits you.",
                "Concrete example": "Given your greek toes, rather avoid very pointy shoes. These might squeeze your toe tips, even if the width of the shoe fits you.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.3", Trigger="arch_length_class == 'long arch' (soft-classified, ±0.005 of tertile boundary)",
             **{"Sentence template": "Given your long arch, the ball of your foot may be pushed into the toe box. A squeezed forefoot may be caused by this instead of actual shoe width, so look for rather short toe boxes.",
                "Concrete example": "Given your long arch, the ball of your foot may be pushed into the toe box. A squeezed forefoot may be caused by this instead of actual shoe width, so look for rather short toe boxes.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "Soft-class fix: 0.733 (raw 'normal') promoted to 'long arch' since within 0.005 of hi=0.734.", "ROMAN COMMENT": ""}),
        dict(ID="T4.4", Trigger="arch_length_class == 'short arch'",
             **{"Sentence template": "Your arch is rather short, meaning your toes are relatively long. Especially when considering aggressive shoes look for sufficient height in the toe box to let your toes curl up.",
                "Concrete example": "Your arch is rather short, meaning your toes are relatively long. Especially when considering aggressive shoes look for sufficient height in the toe box to let your toes curl up.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.5", Trigger="instep_height_class == 'low instep'",
             **{"Sentence template": "Additionally your instep is rather low, so an adjustable closure is preferable to avoid dead space. Ideally double velcro or laces, rather avoid pure slippers.",
                "Concrete example": "Additionally your instep is rather low, so an adjustable closure is preferable to avoid dead space. Ideally double velcro or laces, rather avoid pure slippers.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "Carries the closure recommendation. §3 has a complementary shopping-shorthand sentence (S3.9e).", "ROMAN COMMENT": ""}),
        dict(ID="T4.6", Trigger="instep_height_class == 'high instep'",
             **{"Sentence template": "Additionally your instep is rather high, so an adjustable closure is preferable to actually get into the shoe. Ideally double velcro or laces, you may struggle getting into slippers if adequately downsized.",
                "Concrete example": "Additionally your instep is rather high, so an adjustable closure is preferable to actually get into the shoe. Ideally double velcro or laces, you may struggle getting into slippers if adequately downsized.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.7", Trigger="heel_depth_class == 'deep heel'",
             **{"Sentence template": "Your deep heel projects further back than most, so a deeper, more sculpted heel cup fits naturally.",
                "Concrete example": "Your deep heel projects further back than most, so a deeper, more sculpted heel cup fits naturally.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.8", Trigger="heel_depth_class == 'shallow heel'",
             **{"Sentence template": "Your shallow heel doesn't project as far back as most. Deeply sculpted cups will feel empty at the back.",
                "Concrete example": "Your shallow heel doesn't project as far back as most. Deeply sculpted cups will feel empty at the back.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.9", Trigger="hallux_valgus_class IN ('mild', 'pronounced') AND toe == Egyptian",
             **{"Sentence template": "For your {hva} hallux valgus, we avoid very asymmetric lasts and prefer moderately asymmetric or wider-tipped designs that help the big toe stay in contact with the shoe tip.",
                "Concrete example": "For your mild hallux valgus, we avoid very asymmetric lasts and prefer moderately asymmetric or wider-tipped designs that help the big toe stay in contact with the shoe tip.",
                "Mentions shoe?": "No", "Origin": "V1 verbatim, slotted into V2 trigger", "Status": "LOCKED",
                "Notes": "Verbatim from v1 _para_closure_hva. HVA threshold change (0.25→0.28) deferred to v2 go-live.", "ROMAN COMMENT": ""}),
        dict(ID="T4.10", Trigger="hallux_valgus_class IN ('mild', 'pronounced') AND toe IN (Greek, Roman)",
             **{"Sentence template": "For your {hva} hallux valgus, we prefer a slightly wider, less pointed toe box to compensate for the inward drift.",
                "Concrete example": "For your pronounced hallux valgus, we prefer a slightly wider, less pointed toe box to compensate for the inward drift.",
                "Mentions shoe?": "No", "Origin": "V1 verbatim, slotted into V2 trigger", "Status": "LOCKED",
                "Notes": "Roman + Greek treated the same per Q2.", "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── §2 sheet ───────────────────────────────────────────────────
def build_section2(wb):
    ws = wb.create_sheet("§2 Shoe Fit")
    write_header(ws, "Section 2 — What Your Current Shoe Fit Tells Us",
                 "S1 LOCKED. S2/S3/S4/S5 PROPOSED — final lockdown pending. Re-uses scan-aware V2 helpers (heel/toes/forefoot causes).")

    rows = [
        # S1 — Sizing intro
        dict(ID="S1.1.a", Trigger="N == 1 with size_eu",
             **{"Sentence template": "You wear your {brand} {model} in EU {size_eu}, {downsize_clause}.",
                "Concrete example": "You wear your Scarpa Furia Air in EU 44, 1.5 sizes down from your street size of 45.5.",
                "Mentions shoe?": "Yes", "Origin": "V1 wording, V2 LOCKED", "Status": "LOCKED",
                "Notes": "downsize_clause: '{X} sizes down from your street size of {street}' / 'at your street size of {street}' / '{X} sizes above your street size of {street}'.", "ROMAN COMMENT": ""}),
        dict(ID="S1.1.b", Trigger="N == 1 (always after S1.1.a)",
             **{"Sentence template": "Typical: 'That is a typical fit for {brand}.' OR Other: 'That is {label} for {brand}, where the typical downsize is about {typical_label}.'",
                "Concrete example": "That is a typical fit for Scarpa.\\nOR\\nThat is rather aggressive for La Sportiva, where the typical downsize is about 1 size down.",
                "Mentions shoe?": "Yes (brand only)", "Origin": "V1 wording, V2 LOCKED", "Status": "LOCKED",
                "Notes": "label ∈ {rather aggressive, very aggressive, rather relaxed, very relaxed}.", "ROMAN COMMENT": ""}),
        dict(ID="S1.1.c", Trigger="S1.1.b label is rather/very relaxed AND fit.heel ∈ {empty, loose}",
             **{"Sentence template": "Downsizing further toward the typical range could tighten the heel and reduce the empty feel.",
                "Concrete example": "(appended after S1.1.b)",
                "Mentions shoe?": "No", "Origin": "V1 wording, V2 LOCKED", "Status": "LOCKED",
                "Notes": "Conditional 3rd sentence connecting downsize to heel issue.", "ROMAN COMMENT": ""}),
        dict(ID="S1.2", Trigger="N > 1, all same brand",
             **{"Sentence template": "You wear your {brand} shoes in EU {min}–{max}, {raw_range_clause} from your street size of {street}. For {brand}, where the typical downsize is about {typical_label}, this is a {overall_label} fit overall.",
                "Concrete example": "You wear your Scarpa shoes in EU 43–44.5, 1–2.5 sizes down from your street size of 45.5. For Scarpa, where the typical downsize is about 1.5 sizes down, this is aggressive overall.",
                "Mentions shoe?": "Yes (brand)", "Origin": "V1 wording, V2 LOCKED", "Status": "LOCKED",
                "Notes": "En-dash for ranges OK. raw_range_clause: '1–1.5 sizes down' or single value if min==max.", "ROMAN COMMENT": ""}),
        dict(ID="S1.3.a", Trigger="N > 1, mixed brands, ≥1 deviating shoe",
             **{"Sentence template": "Your shoes range from EU {min} to {max} relative to a street size of {street}. Raw sizes alone don't tell the story because brands size differently: {highlight_1}{connector}{highlight_2}. {overall_clause if N≥3 and aligned}",
                "Concrete example": "Your shoes range from EU 42 to 44 relative to a street size of 45.5. Raw sizes alone don't tell the story because brands size differently: your Scarpa Drago in EU 42 runs aggressive for Scarpa (typical downsize about 1.5 sizes down); also your La Sportiva Theory in EU 43 runs aggressive for La Sportiva (typical downsize about 1.5 sizes down). Across your shoes you tend to size more aggressively than the brand-typical fit.",
                "Mentions shoe?": "Yes (per highlight)", "Origin": "V1 wording + V2 connector logic", "Status": "LOCKED",
                "Notes": "connector: '; also ' when both highlights same direction; ', while ' when opposite or one is typical. overall_clause only N≥3 AND avg aligns with highlights.", "ROMAN COMMENT": ""}),
        dict(ID="S1.3.b", Trigger="N > 1, all shoes near brand-typical (no deviating shoe)",
             **{"Sentence template": "Your shoes range from EU {min} to {max} (street size {street}). Each one sits close to its brand-typical downsize even though the raw EU numbers differ.",
                "Concrete example": "Your shoes range from EU 43 to 45 (street size 45.5). Each one sits close to its brand-typical downsize even though the raw EU numbers differ.",
                "Mentions shoe?": "No", "Origin": "V1 wording, V2 LOCKED", "Status": "LOCKED",
                "Notes": "", "ROMAN COMMENT": ""}),

        # S2 — Fit issues per dimension
        dict(ID="S2.HEEL.empty.1", Trigger="N == 1, heel == empty",
             **{"Sentence template": "Your {brand} {model}'s heel feels empty. {scan_aware_cause}.{ slipper_closure_note if shoe is slipper}",
                "Concrete example": "Your Scarpa Furia Air's heel feels empty. Given your heel is deep but narrow, the cup is likely too wide rather than too deeply sculpted. On top, the slipper has no closure to tighten the heel; a lace-up or two-strap velcro design could compensate.",
                "Mentions shoe?": "Yes", "Origin": "V2 sandbox (scan-aware, vendored)", "Status": "PROPOSED",
                "Notes": "scan_cause from _heel_empty_cause(profile). slipper_closure_note from _heel_empty_closure_note(shoe).", "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.empty.allN", Trigger="N > 1, ALL heel == empty",
             **{"Sentence template": "Your heel feels empty in all your shoes ({Shoe A}, {Shoe B}, …). {scan_aware_cause}.",
                "Concrete example": "Your heel feels empty in all your shoes (Scarpa Furia Air, La Sportiva Solution). Given your heel is deep but narrow, the cup is likely too wide rather than too deeply sculpted.",
                "Mentions shoe?": "Yes (all)", "Origin": "PROPOSED (was generic 'in both shoes')", "Status": "PROPOSED",
                "Notes": "Q1: should we always list shoes by name parenthetically, or use 'all your shoes' generically?", "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.empty.minority", Trigger="N > 1, SOME heel == empty (not all, not contradiction)",
             **{"Sentence template": "Your {brand} {model}'s heel feels empty while your other shoes fit. {scan_aware_cause if profile-decisive, else 'this model's cup runs differently'}.",
                "Concrete example": "Your Scarpa Furia Air's heel feels empty while your other shoes fit. Given your heel is deep but narrow, the cup is likely too wide rather than too deeply sculpted.",
                "Mentions shoe?": "Yes (affected)", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "Q2: should we name the OTHER shoes too, or only the affected?", "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.tight.1", Trigger="N == 1, heel == tight",
             **{"Sentence template": "Your {brand} {model}'s heel feels tight. {scan_aware_cause}.",
                "Concrete example": "Your La Sportiva Solution's heel feels tight. Given your heel is wide and deep, the cup likely runs small in both dimensions.",
                "Mentions shoe?": "Yes", "Origin": "V2 sandbox", "Status": "PROPOSED",
                "Notes": "scan_cause from _heel_tight_cause(profile, shoe).", "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.tight.allN", Trigger="N > 1, ALL heel == tight",
             **{"Sentence template": "Your heel feels tight in all your shoes ({Shoe A}, {Shoe B}, …). {scan_aware_cause}.",
                "Concrete example": "Your heel feels tight in all your shoes (Scarpa Drago, La Sportiva Theory). Given your heel is wide, likely the width is the issue.",
                "Mentions shoe?": "Yes (all)", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.tight.minority", Trigger="N > 1, SOME heel == tight",
             **{"Sentence template": "Your {brand} {model}'s heel feels tight while your other shoes fit.",
                "Concrete example": "Your La Sportiva Solution's heel feels tight while your other shoes fit.",
                "Mentions shoe?": "Yes", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.contradiction", Trigger="N > 1, BOTH empty AND tight ratings present",
             **{"Sentence template": "Your heel fit is inconsistent: empty in your {Shoe A}, tight in your {Shoe B}. Different brands use different heel cup shapes.",
                "Concrete example": "Your heel fit is inconsistent: empty in your Scarpa Furia Air, tight in your La Sportiva Solution. Different brands use different heel cup shapes.",
                "Mentions shoe?": "Yes (both)", "Origin": "V1 wording, V2 sandbox", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),

        dict(ID="S2.TOES.squeezed.1", Trigger="N == 1, toes == squeezed",
             **{"Sentence template": "Your {brand} {model}'s toes feel squeezed. {scan_aware_cause: toe-form mismatch / HVA / wide forefoot / aggressive downsize}.",
                "Concrete example": "Your Scarpa Drago's toes feel squeezed. Your scan shows an Egyptian toe shape while this shoe is built on a Greek toe form, so the most likely cause is a toe-shape mismatch.",
                "Mentions shoe?": "Yes", "Origin": "V2 sandbox", "Status": "PROPOSED",
                "Notes": "scan_cause from _toes_squeezed_cause(profile, shoe).", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.squeezed.allN", Trigger="N > 1, ALL toes == squeezed",
             **{"Sentence template": "Your toes feel squeezed in all your shoes ({Shoe A}, {Shoe B}, …). A consistent squeeze across different lasts points to a structural cause: {scan_cause}.",
                "Concrete example": "Your toes feel squeezed in all your shoes (Scarpa Drago, La Sportiva Solution). A consistent squeeze across different lasts points to a structural cause: your wide forefoot in narrower lasts is the most likely fit.",
                "Mentions shoe?": "Yes (all)", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.squeezed.minority", Trigger="N > 1, SOME toes == squeezed",
             **{"Sentence template": "Your {brand} {model}'s toes feel squeezed while your other shoes fit.",
                "Concrete example": "Your Scarpa Drago's toes feel squeezed while your other shoes fit.",
                "Mentions shoe?": "Yes", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.roomy.1", Trigger="N == 1, toes == roomy",
             **{"Sentence template": "Your {brand} {model}'s toes have extra room. {scan_aware_cause: toe-form mismatch — pick a shoe built on your {user_form} foot}.",
                "Concrete example": "Your La Sportiva Tarantulace's toes have extra room. Your scan shows a Roman toe shape while this shoe is built on an Egyptian toe form, leaving a gap at your second toe.",
                "Mentions shoe?": "Yes", "Origin": "V2 sandbox", "Status": "PROPOSED",
                "Notes": "scan_cause from _toes_roomy_cause(profile, shoe).", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.roomy.allN", Trigger="N > 1, ALL toes == roomy",
             **{"Sentence template": "Your toes have extra room in all your shoes ({Shoe A}, {Shoe B}, …).",
                "Concrete example": "Your toes have extra room in all your shoes (La Sportiva Tarantulace, Scarpa Origin).",
                "Mentions shoe?": "Yes (all)", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.roomy.minority", Trigger="N > 1, SOME toes == roomy",
             **{"Sentence template": "Your {brand} {model}'s toes have extra room while your other shoes fit.",
                "Concrete example": "Your La Sportiva Tarantulace's toes have extra room while your other shoes fit.",
                "Mentions shoe?": "Yes", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.contradiction", Trigger="N > 1, BOTH squeezed AND roomy ratings present",
             **{"Sentence template": "Your toe fit varies: squeezed in your {Shoe A}, roomy in your {Shoe B}. Different toe box shapes; the next section will narrow which one matches your scan.",
                "Concrete example": "Your toe fit varies: squeezed in your Scarpa Drago, roomy in your La Sportiva Tarantulace. Different toe box shapes; the next section will narrow which one matches your scan.",
                "Mentions shoe?": "Yes (both)", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),

        dict(ID="S2.FF.tight.1", Trigger="N == 1, forefoot == tight",
             **{"Sentence template": "Your {brand} {model}'s forefoot feels tight. {scan_aware_cause: lateral width vs. long-arch ball-push}.",
                "Concrete example": "Your Scarpa Drago's forefoot feels tight. Your wide forefoot in this narrower last is the most likely cause.",
                "Mentions shoe?": "Yes", "Origin": "V2 sandbox", "Status": "PROPOSED",
                "Notes": "scan_cause from _forefoot_tight_cause(profile, shoe).", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.tight.allN", Trigger="N > 1, ALL forefoot == tight",
             **{"Sentence template": "Your forefoot feels tight in all your shoes ({Shoe A}, {Shoe B}, …). {scan_aware_cause}.",
                "Concrete example": "Your forefoot feels tight in all your shoes (Scarpa Drago, La Sportiva Solution). Your wide forefoot needs more room than these lasts provide.",
                "Mentions shoe?": "Yes (all)", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.tight.minority", Trigger="N > 1, SOME forefoot == tight",
             **{"Sentence template": "Your {brand} {model}'s forefoot feels tight while your other shoes fit.",
                "Concrete example": "Your Scarpa Drago's forefoot feels tight while your other shoes fit.",
                "Mentions shoe?": "Yes", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.loose.1", Trigger="N == 1, forefoot == loose",
             **{"Sentence template": "Your {brand} {model}'s forefoot feels loose, meaning more volume than your foot needs there.",
                "Concrete example": "Your Five Ten Anasazi's forefoot feels loose, meaning more volume than your foot needs there.",
                "Mentions shoe?": "Yes", "Origin": "V1 wording, V2 PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.loose.allN", Trigger="N > 1, ALL forefoot == loose",
             **{"Sentence template": "Your forefoot feels loose in all your shoes ({Shoe A}, {Shoe B}, …).",
                "Concrete example": "Your forefoot feels loose in all your shoes (Five Ten Anasazi, Evolv Defy).",
                "Mentions shoe?": "Yes (all)", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.loose.minority", Trigger="N > 1, SOME forefoot == loose",
             **{"Sentence template": "Your {brand} {model}'s forefoot feels loose while your other shoes fit.",
                "Concrete example": "Your Five Ten Anasazi's forefoot feels loose while your other shoes fit.",
                "Mentions shoe?": "Yes", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.contradiction", Trigger="N > 1, BOTH tight AND loose forefoot ratings",
             **{"Sentence template": "Your forefoot fit is split: tight in your {Shoe A}, loose in your {Shoe B}. Different lasts.",
                "Concrete example": "Your forefoot fit is split: tight in your Scarpa Drago, loose in your Five Ten Anasazi. Different lasts.",
                "Mentions shoe?": "Yes (both)", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),

        # S3 — Cross-shoe insights
        dict(ID="S3.BRAND_INCONSISTENCY", Trigger="N ≥ 2, ≥2 same brand, downsize spread ≥0.75 sizes, divergent fits",
             **{"Sentence template": "Within {brand}, your {good_model} at EU {good_size} fits well, while your {bad_model} at EU {bad_size} has {issues}. The {spread}-size spread indicates {brand}'s lasts vary by model; brand alone isn't a reliable predictor of fit for you.",
                "Concrete example": "Within La Sportiva, your Solution at EU 43 fits well, while your Tarantulace at EU 44.5 has loose forefoot and empty heel. The 1.5-size spread indicates La Sportiva's lasts vary by model; brand alone isn't a reliable predictor of fit for you.",
                "Mentions shoe?": "Yes (both)", "Origin": "V2 sandbox NEW", "Status": "PROPOSED",
                "Notes": "Steps 1-7 work, currently lives only in interp_shoe_fit_v2.py.", "ROMAN COMMENT": ""}),
        dict(ID="S3.HEEL_VOL_VS_SHAPE", Trigger="N ≥ 2, one heel == perfect AND another heel == empty AND BOTH same db_heel_volume",
             **{"Sentence template": "Your {good_model} fits perfectly in the heel while your {bad_model} feels empty, even though both are rated {volume} heel volume. The difference is cup shape, not volume.",
                "Concrete example": "Your Skwama fits perfectly in the heel while your Solution feels empty, even though both are rated medium heel volume. The difference is cup shape, not volume.",
                "Mentions shoe?": "Yes (both)", "Origin": "V2 sandbox NEW", "Status": "PROPOSED",
                "Notes": "Useful insight when DB volume rating doesn't predict fit.", "ROMAN COMMENT": ""}),
        dict(ID="S3.HEEL_SOFT_CONFORM", Trigger="N ≥ 1, soft shoe (db_stiffness < 0.4) + heel == perfect + scan extreme heel that contradicts shoe's db_heel_volume",
             **{"Sentence template": "Your {brand} {model}'s heel fits despite a mismatch between the cup volume and your scan. The shoe is soft enough to conform; in a stiffer model with the same cup, the same fit might not work.",
                "Concrete example": "Your Scarpa Furia Air's heel fits despite a mismatch between the cup volume and your scan. The shoe is soft enough to conform; in a stiffer model with the same cup, the same fit might not work.",
                "Mentions shoe?": "Yes", "Origin": "V2 sandbox NEW", "Status": "PROPOSED",
                "Notes": "Warns user about projection risk when shopping stiffer.", "ROMAN COMMENT": ""}),
        dict(ID="S3.TOE_FORM_SOFT_MASK", Trigger="N ≥ 1, user toe ≠ shoe db_toe_form + soft shoe + toes==perfect + forefoot==perfect",
             **{"Sentence template": "Your {brand} {model} is built on a {shoe_form} toe form, yet your {user_form} toes feel fine. The shoe is soft enough to conform; stiffer shoes on the same form may not.",
                "Concrete example": "Your Scarpa Furia Air is built on a Greek toe form, yet your Egyptian toes feel fine. The shoe is soft enough to conform; stiffer shoes on the same form may not.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "PROPOSED",
                "Notes": "From V1 _para_toe_form_mismatch.", "ROMAN COMMENT": ""}),
        dict(ID="S3.SHALLOW_HEEL_EXCEPTION", Trigger="N ≥ 1, profile.heel_depth==shallow + ≥1 shoe with heel==perfect",
             **{"Sentence template": "Your shallow heel typically causes empty feel in standard cups, yet your {brand} {model} (rated {volume} heel volume) fits. Use this as the heel-cup reference for new picks.",
                "Concrete example": "Your shallow heel typically causes empty feel in standard cups, yet your Scarpa Vapor V (rated narrow heel volume) fits. Use this as the heel-cup reference for new picks.",
                "Mentions shoe?": "Yes (exception shoe)", "Origin": "V1 + V2", "Status": "PROPOSED",
                "Notes": "From V1 _para_heel_depth_insight.", "ROMAN COMMENT": ""}),

        # S4 — Combined squeeze+loose
        dict(ID="S4.SQUEEZE_LOOSE", Trigger="Same shoe: toes == squeezed AND forefoot ∈ {loose, roomy}",
             **{"Sentence template": "Your {brand} {model} shows an unusual combination: toes squeezed while the forefoot is loose. The toe form is the mismatch, not the width — {scan_cause}.",
                "Concrete example": "Your Scarpa Drago shows an unusual combination: toes squeezed while the forefoot is loose. The toe form is the mismatch, not the width — your scan shows Egyptian toes while this shoe is built on a Greek toe form.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "PROPOSED",
                "Notes": "When S4 fires for a shoe, suppress S2.TOES.squeezed AND S2.FOREFOOT.loose for that same shoe. Q4: 'The toe form is the mismatch, not the width' — confirm OK or rephrase.", "ROMAN COMMENT": ""}),

        # S5 — Closing anchor
        dict(ID="S5.PERFECT_ALL", Trigger="≥1 shoe with perfect on all 3 dims",
             **{"Sentence template": "We use your {brand} {model} as a reliable reference and target shoes with similar fit characteristics.",
                "Concrete example": "We use your Scarpa Vapor V as a reliable reference and target shoes with similar fit characteristics.",
                "Mentions shoe?": "Yes", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S5.PARTIAL", Trigger="≥1 shoe with perfect on ≥1 dim AND imperfect on ≥1 dim",
             **{"Sentence template": "We use your {best_model} as anchor and focus on closing the gaps where it falls short ({imperfect_list}).",
                "Concrete example": "We use your Scarpa Furia Air as anchor and focus on closing the gaps where it falls short (heel feels empty).",
                "Mentions shoe?": "Yes", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S5.NONE_GOOD", Trigger="No shoe has any perfect dim",
             **{"Sentence template": "None of your shoes is a clean reference. We'll lean more on the scan than on shoe-fit feedback for the recommendation.",
                "Concrete example": "None of your shoes is a clean reference. We'll lean more on the scan than on shoe-fit feedback for the recommendation.",
                "Mentions shoe?": "No", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "Q3: is this voice OK or dismissive of the user's input?", "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── §3 sheet ───────────────────────────────────────────────────
def build_section3(wb):
    ws = wb.create_sheet("§3 What to Look For")
    write_header(ws, "Section 3 — What to Look For",
                 "Mix of V1-LIVE (currently emitted by /scan), V2-DRAFT (sandbox), DROPPED (V1 paragraphs V2 removes), and proposed cleanups.")

    rows = [
        # Target profile (always)
        dict(ID="S3.1", Trigger="Always; scan synthesis with fit feedback",
             **{"Sentence template": "Based on your scan, we target {meas_fw_label}-width shoes with {meas_hv_label} heel volume.",
                "Concrete example": "Based on your scan, we target narrow-width shoes with narrow heel volume.",
                "Mentions shoe?": "No", "Origin": "V1 + V2 (unchanged)", "Status": "V1-LIVE",
                "Notes": "5+ variants based on scan/feedback agreement. See _para_target_profile.", "ROMAN COMMENT": ""}),
        dict(ID="S3.2a", Trigger="Single perfect-fit shoe diverges from scan on both width + heel",
             **{"Sentence template": "Your {brand} {model} ({shoe_w} width, {shoe_hv} heel volume) fits perfectly despite your scan suggesting {meas_fw_label}. We follow your shoe's proven fit and target {target_fw_label} width with {target_hv_label} heel volume.",
                "Concrete example": "Your Scarpa Vapor V (medium width, narrow heel volume) fits perfectly despite your scan suggesting wide. We follow your shoe's proven fit and target medium width with narrow heel volume.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.2b", Trigger="Scan + feedback agree; shoes fit perfectly",
             **{"Sentence template": "Your {brand} {model} ({shoe_w} width) confirms this.",
                "Concrete example": "Your Scarpa Vapor V (narrow width) confirms this.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.2c", Trigger="Feedback shifts target away from scan",
             **{"Sentence template": "Your scan points to {meas_fw_label}-width shoes with {meas_hv_label} heel volume, but your shoe feedback adjusts both.",
                "Concrete example": "Your scan points to wide-width shoes with normal heel volume, but your shoe feedback adjusts both.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # Forefoot width explainer variants
        dict(ID="S3.3a", Trigger="≥2 shoes tight in forefoot; clamped",
             **{"Sentence template": "In {count_word} shoes the forefoot or toes feel too tight. This is likely shoe-specific (toe box shape mismatch) rather than your {meas_label} foot needing much wider shoes. We go one level wider to {target_label}.",
                "Concrete example": "In both shoes the forefoot or toes feel too tight. This is likely shoe-specific (toe box shape mismatch) rather than your normal foot needing much wider shoes. We go one level wider to wide.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.3c", Trigger="single tight shoe; clamped",
             **{"Sentence template": "Your {brand} {model} ({shoe_w} width) {fit_issue_desc}. This is likely a shoe-specific fit issue (toe box shape or last geometry) rather than your {meas_label} foot needing a much wider shoe. We go one level wider to {target_label}.",
                "Concrete example": "Your Scarpa Drago (narrow width) feels tight in the forefoot. This is likely a shoe-specific fit issue (toe box shape or last geometry) rather than your normal foot needing a much wider shoe. We go one level wider to medium.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.3d", Trigger="loose forefoot, wider shoe than scan; target lands between",
             **{"Sentence template": "Your {brand} {model} ({shoe_w} width) feels loose in the forefoot, confirming you don't need that much room. We target {target_label} width, narrower than the shoe but wider than your scan's {meas_label} measurement.",
                "Concrete example": "Your Five Ten Anasazi (wide width) feels loose in the forefoot, confirming you don't need that much room. We target medium width, narrower than the shoe but wider than your scan's narrow measurement.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.3f", Trigger="perfect fit in different width; user has preference",
             **{"Sentence template": "Your {brand} {model} ({shoe_w} width) fits perfectly despite your scan suggesting {meas_label}. You prefer a {direction} forefoot, so we target {target_label} width.",
                "Concrete example": "Your Scarpa Vapor V (medium width) fits perfectly despite your scan suggesting wide. You prefer a narrower forefoot, so we target medium width.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # Heel volume explainer variants
        dict(ID="S3.4c", Trigger="single empty heel, clamped",
             **{"Sentence template": "Your heel feels empty in the {brand} {model} ({shoe_hv} heel volume). This is likely shoe-specific (heel cup geometry) rather than your {meas_label} heel needing much tighter shoes. We go one level tighter to {target_label}.",
                "Concrete example": "Your heel feels empty in the Scarpa Furia Air (medium heel volume). This is likely shoe-specific (heel cup geometry) rather than your narrow heel needing much tighter shoes. We go one level tighter to narrow.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.4d", Trigger="tight heel; widen target",
             **{"Sentence template": "Your heel feels tight in the {brand} {model} ({shoe_hv} heel volume), so we widen the heel target from {meas_label} to {target_label}.",
                "Concrete example": "Your heel feels tight in the La Sportiva Solution (narrow heel volume), so we widen the heel target from narrow to medium.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # Stiffness adjustment paragraphs
        dict(ID="S3.5a", Trigger="divergence + soft shoe + tight forefoot",
             **{"Sentence template": "Keep in mind that your {brand} {model} is a {feel} shoe. Soft shoes stretch and mold to the foot, making a snug forefoot more forgiving than the same width in a stiff shoe. For the stiffer recommendations below, we compensate by targeting one level wider.",
                "Concrete example": "Keep in mind that your Scarpa Drago is a sensitive shoe. Soft shoes stretch and mold to the foot, making a snug forefoot more forgiving than the same width in a stiff shoe. For the stiffer recommendations below, we compensate by targeting one level wider.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "Cross-tier stiffness adjustment.", "ROMAN COMMENT": ""}),
        # V1 preference paragraphs (DROPPED in V2)
        dict(ID="S3.6 (all)", Trigger="next_shoe_preference IS NOT NULL (any of comfort/performance/same/softer/stiffer/etc)",
             **{"Sentence template": "(17 distinct variants — see V1 _para_preference) e.g. 'You are looking for a performance shoe...', 'You want a shoe similar to your {brand} {model} but softer.', 'You mentioned wanting a stiffer shoe...'",
                "Concrete example": "You want a shoe similar to your Scarpa Furia Air but stiffer. We match the geometry and fit profile while leaning toward higher stiffness across all tiers.",
                "Mentions shoe?": "Sometimes", "Origin": "V1 only", "Status": "DROPPED",
                "Notes": "All 17 variants DROPPED in V2 because we removed the next_shoe_preference input. The aggressiveness wizard step partly replaces it; explicit preference no longer captured.", "ROMAN COMMENT": ""}),
        # Forefoot paradox
        dict(ID="S3.7", Trigger="toes squeezed + forefoot loose + long arch",
             **{"Sentence template": "Despite your wide scan measurement, the forefoot feels loose in your {brand} {model}. Your long arch shifts the ball of your foot backward... You need one where the forefoot volume sits further back, matching your proportions.",
                "Concrete example": "Despite your wide scan measurement, the forefoot feels loose in your Scarpa Drago. Your long arch shifts the ball of your foot backward... You need one where the forefoot volume sits further back, matching your proportions.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # Toe form guidance
        dict(ID="S3.8a", Trigger="Toe shape mismatch + toes squeezed + a confirming shoe exists",
             **{"Sentence template": "Your {squeezed_brand} {squeezed_model} has a {shoe_label} toe form that squeezes your {user_label} toes, while your {user_label}-compatible {matched_brand} {matched_model} fits fine. For all recommendations, we filter for {user_label}-compatible toe boxes.",
                "Concrete example": "Your Scarpa Drago has a Greek toe form that squeezes your Egyptian toes, while your Egyptian-compatible Scarpa Vapor V fits fine. For all recommendations, we filter for Egyptian-compatible toe boxes.",
                "Mentions shoe?": "Yes (both)", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Strongest toe-form signal.", "ROMAN COMMENT": ""}),
        # Closure / HVA  V1 vs V2
        dict(ID="S3.9a", Trigger="High instep (V1)",
             **{"Sentence template": "With your high instep, lace-up closures are ideal because they open up over the midfoot without affecting heel or forefoot fit. Slippers and single-strap velcros tend to press on high insteps.",
                "Concrete example": "(same)",
                "Mentions shoe?": "No", "Origin": "V1 only", "Status": "DROPPED",
                "Notes": "DROPPED in V2 — replaced by S3.9e + T4.6 in §1 carries the recommendation.", "ROMAN COMMENT": ""}),
        dict(ID="S3.9b", Trigger="Low instep (V1)",
             **{"Sentence template": "With your low instep, slippers and velcro closures work well. Lace-ups also work if you cinch them down across the instep to avoid gapping.",
                "Concrete example": "(same)",
                "Mentions shoe?": "No", "Origin": "V1 only", "Status": "DROPPED",
                "Notes": "DROPPED in V2 — wrong rule (slippers actually struggle for low instep) AND used banned word 'cinch'.", "ROMAN COMMENT": ""}),
        dict(ID="S3.9e", Trigger="instep_height_class IN ('low instep', 'high instep')",
             **{"Sentence template": "In the shoe recommendations we avoid pure slippers, as these might create issues given your {label} instep.",
                "Concrete example": "In the shoe recommendations we avoid pure slippers, as these might create issues given your low instep.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "Replaces V1 closure paragraphs. Complements §1 T4.5/T4.6.", "ROMAN COMMENT": ""}),
        dict(ID="S3.9f", Trigger="HVA IN ('mild', 'pronounced') AND toe == Egyptian",
             **{"Sentence template": "For your {hva} hallux valgus, we avoid very asymmetric lasts and prefer moderately asymmetric or wider-tipped designs that help the big toe stay in contact with the shoe tip.",
                "Concrete example": "For your mild hallux valgus, we avoid very asymmetric lasts and prefer moderately asymmetric or wider-tipped designs that help the big toe stay in contact with the shoe tip.",
                "Mentions shoe?": "No", "Origin": "V2 (verbatim from V1)", "Status": "V2-DRAFT",
                "Notes": "Same wording also lives in §1 T4.9. Possible duplication — flag if §3 should drop this.", "ROMAN COMMENT": ""}),
        dict(ID="S3.9g", Trigger="HVA IN ('mild', 'pronounced') AND toe IN (Greek, Roman)",
             **{"Sentence template": "For your {hva} hallux valgus, we prefer a slightly wider, less pointed toe box to compensate for the inward drift.",
                "Concrete example": "For your pronounced hallux valgus, we prefer a slightly wider, less pointed toe box to compensate for the inward drift.",
                "Mentions shoe?": "No", "Origin": "V2 (verbatim from V1)", "Status": "V2-DRAFT",
                "Notes": "Same wording in §1 T4.10. Possible duplication.", "ROMAN COMMENT": ""}),
        # Tradeoffs (general)
        dict(ID="S3.10a", Trigger="wide forefoot + narrow heel; no feedback divergence",
             **{"Sentence template": "When the forefoot and heel need opposite things, we prioritize forefoot fit. A loose heel is manageable with technique, but a squeezed forefoot causes pain.",
                "Concrete example": "(same)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.10c", Trigger="High instep AND user wears slipper",
             **{"Sentence template": "You wear a slipper despite a high instep. We include both lace-up alternatives and slipper options since you are accustomed to the style.",
                "Concrete example": "(same)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.10d", Trigger="Contradictory heel feedback across shoes",
             **{"Sentence template": "Your heel feedback is contradictory (empty in some shoes, tight in others). This reflects different heel cup shapes rather than a universal issue. We aim for the middle ground, prioritizing shoes closest to where the heel fits best.",
                "Concrete example": "(same)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # Fit context
        dict(ID="S3.11e", Trigger="One shoe fits notably better than others",
             **{"Sentence template": "Among your shoes, the {brand} {model} provides the best overall fit. We weight its geometry most heavily in the recommendations.",
                "Concrete example": "Among your shoes, the Scarpa Vapor V provides the best overall fit. We weight its geometry most heavily in the recommendations.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # Shallow heel
        dict(ID="S3.12c", Trigger="Shallow heel + empty heels in many shoes; no exception (V1)",
             **{"Sentence template": "Your shallow heel profile means most heel cups feel empty because they are too deeply sculpted for your heel shape. We target narrow heel volume to minimize the gap.",
                "Concrete example": "(same)",
                "Mentions shoe?": "No", "Origin": "V1 only", "Status": "DROPPED",
                "Notes": "Dropped in V2 because the 'narrow heel volume workaround' overstates what we can deliver — heel volume ≠ heel depth.", "ROMAN COMMENT": ""}),
        dict(ID="S3.12d", Trigger="Shallow heel profile present (V2 replacement)",
             **{"Sentence template": "A heel cup may feel too deep for your shallow heel. We are working to get heel depth data for shoes so we can give better recommendations moving forward. Right now the suggestions are limited to heel width.",
                "Concrete example": "(same)",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "V2-DRAFT",
                "Notes": "Honest disclaimer instead of overpromise. Roman 2026-04-25.", "ROMAN COMMENT": ""}),
        # V2 NEW — target shape (aggressiveness derived)
        dict(ID="S3.13", Trigger="V2 inputs present: aggressiveness + target dict",
             **{"Sentence template": "For your {intent} in {ctx}, we target {shape}. Stiffness, downturn, and asymmetry all anchor on this choice. The softer, stiffer, and budget tiers below offer adjacent options around the same shape.",
                "Concrete example": "For your aggressive selection in bouldering on sandstone, we target an aggressive downturn with strong asymmetry. Stiffness, downturn, and asymmetry all anchor on this choice. The softer, stiffer, and budget tiers below offer adjacent options around the same shape.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "V2-DRAFT",
                "Notes": "Inserted right after S3.1 P1 target. Bridges aggressiveness wizard input → geometry recommendation.", "ROMAN COMMENT": ""}),
        # V2 NEW — asym adjustment
        dict(ID="S3.14", Trigger="V2 target dict + asym_delta > 0 + baseline_lbl ≠ target_lbl",
             **{"Sentence template": "Your Egyptian toe shape (longest at the big toe) lines up with a slightly more asymmetric last, so we shift the asymmetry target up from {baseline_lbl} to {target_lbl}. The shoe's curve follows your foot's natural axis instead of fighting it.",
                "Concrete example": "Your Egyptian toe shape (longest at the big toe) lines up with a slightly more asymmetric last, so we shift the asymmetry target up from moderate to strong. The shoe's curve follows your foot's natural axis instead of fighting it.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "V2-DRAFT",
                "Notes": "Tautology guard added: skip when baseline_lbl == target_lbl (caught by Roman 2026-04-27).", "ROMAN COMMENT": ""}),
        dict(ID="S3.14b", Trigger="V2 target dict + asym_delta < 0 + HVA present",
             **{"Sentence template": "Your hallux deviation pushes the asymmetry target down from {baseline_lbl} to {target_lbl}. A more symmetric last keeps the big-toe joint from being squeezed inward, where an aggressively asymmetric shoe would aggravate it.",
                "Concrete example": "Your hallux deviation pushes the asymmetry target down from strong to moderate. A more symmetric last keeps the big-toe joint from being squeezed inward, where an aggressively asymmetric shoe would aggravate it.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "V2-DRAFT",
                "Notes": "", "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── Shoe Card P1 ───────────────────────────────────────────────
def build_card_p1(wb):
    ws = wb.create_sheet("Card P1 (description)")
    write_header(ws, "Shoe Card — Paragraph 1 (description)",
                 "What the shoe IS. V2 strips toe-box and downturn-with-asymmetry sentences when they match user target (Roman 2026-04-27 dedup).")

    rows = [
        dict(ID="P1.1", Trigger="width AND heel_volume defined",
             **{"Sentence template": "{Width} fit throughout{vol_suffix}",
                "Concrete example": "Narrow fit throughout, low forefoot volume",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "vol_suffix: ', low/high forefoot volume' or empty.", "ROMAN COMMENT": ""}),
        dict(ID="P1.2", Trigger="width defined, heel_volume absent",
             **{"Sentence template": "{Width} forefoot{vol_suffix}",
                "Concrete example": "Narrow forefoot",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P1.3", Trigger="closure == 'slipper'",
             **{"Sentence template": "Slipper construction",
                "Concrete example": "Narrow fit throughout, slipper construction.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "Appended to fit_str.", "ROMAN COMMENT": ""}),
        dict(ID="P1.4", Trigger="closure == 'lace'",
             **{"Sentence template": "Lace closure",
                "Concrete example": "Medium fit throughout, lace closure.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P1.5", Trigger="toe_form list non-empty (V1 — V2 STRIPS when matches user)",
             **{"Sentence template": "{Toe_forms_joined} toe box",
                "Concrete example": "Egyptian toe box.",
                "Mentions shoe?": "No", "Origin": "V1 (V2 strips matching)", "Status": "V2-DRAFT",
                "Notes": "V2 STRIPS this sentence when user's toe shape is in the shoe's toe_form list. Only kept when shoe deviates.", "ROMAN COMMENT": ""}),
        dict(ID="P1.6+7", Trigger="downturn defined (V1 — V2 STRIPS when matches target)",
             **{"Sentence template": "{Downturn_label} OR {Downturn_label} with {Asymmetry_label}",
                "Concrete example": "Aggressively downturned with strong asymmetry.",
                "Mentions shoe?": "No", "Origin": "V1 (V2 strips matching)", "Status": "V2-DRAFT",
                "Notes": "V2 STRIPS via _P1_PERF_RE regex when sentence matches target_dt + target_asym pair. Only kept when shoe deviates.", "ROMAN COMMENT": ""}),
        dict(ID="P1.8", Trigger="no_edge == True",
             **{"Sentence template": "No-edge design wraps rubber smoothly around the sole for smearing but less precision on small edges",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Always kept (genuine shoe-specific characteristic).", "ROMAN COMMENT": ""}),
        dict(ID="P1.9", Trigger="rubber_type and/or rubber_thickness_mm exist",
             **{"Sentence template": "{thickness}mm {rubber_compound} rubber with a {midsole_coverage}, {midsole_stiffness} midsole, resulting in a {stiffness_word} shoe",
                "Concrete example": "3.5mm Vibram XS Grip rubber with a full, soft midsole, resulting in a sensitive shoe.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Roman locked sole format 2026-04-04.", "ROMAN COMMENT": ""}),
        dict(ID="P1.12", Trigger="stiffness computed exists",
             **{"Sentence template": "{stiffness_word} shoe",
                "Concrete example": "Balanced shoe.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "7 levels: super sensitive, very sensitive, sensitive, balanced, supportive, very supportive, super supportive.", "ROMAN COMMENT": ""}),
        dict(ID="P1.13", Trigger="description contains 'P3' (La Sportiva)",
             **{"Sentence template": "P3 system maintains downturn shape over time.",
                "Concrete example": "(appended)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P1.14", Trigger="special_fit_notes field populated in DB",
             **{"Sentence template": "{special_fit_notes}",
                "Concrete example": "(arbitrary string)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Free-form per-shoe note.", "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── Shoe Card P2 ───────────────────────────────────────────────
def build_card_p2(wb):
    ws = wb.create_sheet("Card P2 (why selected)")
    write_header(ws, "Shoe Card — Paragraph 2 (why selected)",
                 "Why this shoe vs others. V2 drops the 'Good overall fit...' boilerplate fallback (returns None instead).")

    rows = [
        # Reference-shoe feedback reasons
        dict(ID="P2.1", Trigger="heel_empty score ≥3, <70% peers share",
             **{"Sentence template": "the tighter heel should fix the empty-heel feeling from your current shoe",
                "Concrete example": "Selected because the tighter heel should fix the empty-heel feeling from your current shoe.",
                "Mentions shoe?": "No (refers to 'current shoe')", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.2", Trigger="heel_tight score ≥3",
             **{"Sentence template": "the roomier heel relieves the tightness you felt in your current shoe",
                "Concrete example": "Selected because the roomier heel relieves the tightness you felt in your current shoe.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.3", Trigger="toes_squeezed score ≥3",
             **{"Sentence template": "the wider toe box gives your toes more room than your current shoe",
                "Concrete example": "Selected because the wider toe box gives your toes more room than your current shoe.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.4", Trigger="toes_roomy score ≥3",
             **{"Sentence template": "the snugger toe box avoids the loose feeling from your current shoe",
                "Concrete example": "Selected because the snugger toe box avoids the loose feeling from your current shoe.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.5", Trigger="ff_tight score ≥3",
             **{"Sentence template": "the wider forefoot relieves the tightness you felt",
                "Concrete example": "Selected because the wider forefoot relieves the tightness you felt.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.6", Trigger="ff_loose score ≥3",
             **{"Sentence template": "the snugger forefoot fixes the loose feeling you experienced",
                "Concrete example": "Selected because the snugger forefoot fixes the loose feeling you experienced.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # Anatomy match reasons
        dict(ID="P2.7-8", Trigger="shallow_heel score ≥3 (with/without heap)",
             **{"Sentence template": "the narrow heel cup fits your shallow heel and should fix the empty-heel feeling from your current shoes (combined) OR the narrow heel cup suits your shallow heel (standalone)",
                "Concrete example": "Selected because the narrow heel cup fits your shallow heel and should fix the empty-heel feeling from your current shoes.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.9", Trigger="instep_downturn score ≥3",
             **{"Sentence template": "the downturn relieves pressure on your high instep",
                "Concrete example": "Selected because the downturn relieves pressure on your high instep.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # Tier-anchored reasons
        dict(ID="P2.10", Trigger="stiffness differs from peer avg by ±0.10+",
             **{"Sentence template": "noticeably {stiffer|softer} than most other picks, for more {support|sensitivity}",
                "Concrete example": "Selected because noticeably stiffer than most other picks, for more support.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Tier standout — fires often (3 times in Roman's review).", "ROMAN COMMENT": ""}),
        dict(ID="P2.12", Trigger="downturn rank differs from peer avg by >0.8",
             **{"Sentence template": "the {dt_word} downturn is {more|less} aggressive than most picks, {built for steeper terrain|trading power for comfort}",
                "Concrete example": "Selected because the slight downturn is less aggressive than most picks, trading power for comfort.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.14", Trigger="feel differs from peer avg by >1.2 ranks",
             **{"Sentence template": "the {shoe_feel} feel is {softer|firmer} than most other picks",
                "Concrete example": "Selected because the sensitive feel is softer than most other picks.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.16", Trigger="width stands out vs peers (±0.5 rank)",
             **{"Sentence template": "the {wider|narrower} forefoot {gives your toes more room|wraps tighter} than most other picks",
                "Concrete example": "Selected because the wider forefoot gives your toes more room than most other picks.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # No-edge
        dict(ID="P2.17-18", Trigger="no_edge == True (only / shared)",
             **{"Sentence template": "the only edgeless shoe in this set, wrapping around footholds for extra contact (only) OR edgeless design wraps around footholds for extra contact (shared)",
                "Concrete example": "Selected because edgeless design wraps around footholds for extra contact.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # Closure standout
        dict(ID="P2.19", Trigger="closure differs from peer majority",
             **{"Sentence template": "the {lace-up closure|slipper design|velcro closure} {allows custom fit|gives low-profile sensitive fit|allows quick on/off}",
                "Concrete example": "Selected because the lace-up closure allows a more custom fit.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        # Fallback boilerplate (V1 only, V2 returns None)
        dict(ID="P2.23", Trigger="No specific reason fires (V1 fallback)",
             **{"Sentence template": "Good overall fit for your foot shape and climbing style.",
                "Concrete example": "(verbatim, no shoe context)",
                "Mentions shoe?": "No", "Origin": "V1 only", "Status": "DROPPED",
                "Notes": "DROPPED in V2 (returns None instead). Showed up 4/12 times in Roman's review — main P2 deduplication win.", "ROMAN COMMENT": ""}),
        # Stock + price
        dict(ID="P2.24", Trigger="not_in_stock == True",
             **{"Sentence template": "Note: this shoe is not currently available online. Check local shops or wait for restocks.",
                "Concrete example": "(appended to P2)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.25", Trigger="tier == 'budget' AND best_price exists",
             **{"Sentence template": "At EUR {price:.0f}, this is a strong value pick.",
                "Concrete example": "At EUR 124, this is a strong value pick.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Prepended to P2. Slightly redundant with size pill (which already shows EUR price).", "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── Shoe Card P3 ───────────────────────────────────────────────
def build_card_p3(wb):
    ws = wb.create_sheet("Card P3 (tradeoffs)")
    write_header(ws, "Shoe Card — Paragraph 3 (tradeoffs)",
                 "Compromises this shoe makes vs your foot. V2 drops the 'No notable tradeoffs...' fallback (returns None — card omits P3 entirely).")

    rows = [
        dict(ID="P3.2", Trigger="forefoot_width axis < 0 (V2)",
             **{"Sentence template": "the {shoe_width} forefoot runs {wider|narrower} than your {target_label} target",
                "Concrete example": "Tradeoff: The medium forefoot runs wider than your narrow target.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P3.4", Trigger="heel_volume axis < 0 (V2)",
             **{"Sentence template": "the {shoe_hv} heel volume is {roomier|tighter} than your {target_label} target",
                "Concrete example": "Tradeoff: The medium heel volume is roomier than your narrow target.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P3.7", Trigger="toe_form axis < 0 (V2), user_toe not in shoe forms",
             **{"Sentence template": "the toe box is shaped for {shoe_forms} feet, not {user_toe} (Greek case) OR built for {shoe_forms} feet, opposite to your {user_toe} foot (Egyptian↔Roman)",
                "Concrete example": "Tradeoff: The toe box is shaped for Egyptian feet, not Greek.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT",
                "Notes": "V2: Greek mismatch -5×conf, Egyptian/Roman opposites -10×conf.", "ROMAN COMMENT": ""}),
        dict(ID="P3.15", Trigger="downturn axis < 0 (V2), shoe_dt vs target",
             **{"Sentence template": "the {dt_short} downturn is {less|more} aggressive than the {target_label} downturn we target for your selection",
                "Concrete example": "Tradeoff: The slight downturn is less aggressive than the aggressive downturn we target for your selection.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P3.18", Trigger="asymmetry axis < 0 (V2), shoe_as vs target",
             **{"Sentence template": "the {shoe_asym} asymmetry is {more|less} {aggressive|pronounced} than the {target_label} asymmetry we target",
                "Concrete example": "Tradeoff: The slight asymmetry is less pronounced than the strong asymmetry we target.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT",
                "Notes": "Target_asym is foot-shape adjusted (e.g., +1 for Egyptian + no HVA).", "ROMAN COMMENT": ""}),
        dict(ID="P3.20", Trigger="stiffness axis < 0 (V2), tier-aware",
             **{"Sentence template": "the {stiff_word} sole is {softer|stiffer} than what you currently climb in (with anchor) OR sits outside the comfortable range for your selection (no anchor)",
                "Concrete example": "Tradeoff: The supportive sole is stiffer than what you currently climb in.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT",
                "Notes": "Tier-aware: softer tier doesn't surface 'stiffer than current' as a tradeoff; stiffer tier doesn't surface 'softer than current'.", "ROMAN COMMENT": ""}),
        dict(ID="P3.23", Trigger="instep_extreme axis < 0, shoe_closure == 'slipper', user_instep not normal",
             **{"Sentence template": "the slipper closure leaves no adjustability for your {user_instep_clean} instep",
                "Concrete example": "Tradeoff: The slipper closure leaves no adjustability for your low instep.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P3.39", Trigger="discipline_overlap axis < 0 (V2)",
             **{"Sentence template": "the shoe is not built for {discipline} climbing",
                "Concrete example": "Tradeoff: The shoe is not built for trad climbing.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT",
                "Notes": "V2: -100 score, overrides all other reasons. Filters out unsuitable picks before P3 even fires for them.", "ROMAN COMMENT": ""}),
        dict(ID="P3.40", Trigger="closure axis < 0 (V2)",
             **{"Sentence template": "the {shoe_closure} closure is not ideal for {discipline} climbing",
                "Concrete example": "Tradeoff: The slipper closure is not ideal for trad/multipitch climbing.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P3.42", Trigger="No tradeoff issues at all (V1 fallback)",
             **{"Sentence template": "No notable tradeoffs for your foot shape.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V1 only", "Status": "DROPPED",
                "Notes": "DROPPED in V2 (returns None — card omits P3 entirely). Showed up 10/12 times in Roman's review — main P3 deduplication win.", "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── Cross-cutting (sizing rec, tier headers) ──────────────────
def build_cross(wb):
    ws = wb.create_sheet("Cross-cutting")
    write_header(ws, "Cross-cutting — sizing, tier headers, anchors",
                 "Shared elements that appear across all results pages.")

    rows = [
        dict(ID="X.SIZE", Trigger="Per-shoe (every recommendation card)",
             **{"Sentence template": "EU {recommended_size_eu}",
                "Concrete example": "EU 44 (size pill in shoe card header)",
                "Mentions shoe?": "Implicit", "Origin": "V1 + V2", "Status": "LOCKED",
                "Notes": "MUST snap to half-EU via round(x*2)/2. _calc_recommended_size returns 0.1-precision garbage like 45.2 — must be snapped before display (memory: feedback_recommended_size_eu_must_snap).", "ROMAN COMMENT": ""}),
        dict(ID="X.PRICE", Trigger="Budget tier shoes only",
             **{"Sentence template": "{price:.0f} EUR @ {retailer}",
                "Concrete example": "124 EUR @ TheOutlet (in size pill)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="X.TIER.baseline", Trigger="Tier header (always shown if tier has picks)",
             **{"Sentence template": "Your Best Match — Similar feel and use case to your current shoes",
                "Concrete example": "(verbatim header)",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED",
                "Notes": "From CATEGORY_META.", "ROMAN COMMENT": ""}),
        dict(ID="X.TIER.softer", Trigger="Softer tier header",
             **{"Sentence template": "Softer Shoes — For more sensitivity, recommended for indoors and bouldering",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="X.TIER.stiffer", Trigger="Stiffer tier header",
             **{"Sentence template": "Stiffer Shoes — For more support, recommended for outdoors and sport/trad climbing",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="X.TIER.budget", Trigger="Budget tier header",
             **{"Sentence template": "Best Value — Affordable picks at your recommended size",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="X.SEEMORE", Trigger="Per tier (after all picks shown)",
             **{"Sentence template": "See more {top matches|softer picks|stiffer picks|value picks} →",
                "Concrete example": "See more top matches →",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED",
                "Notes": "Links to /scan/{scan_id}/browse?tier={cat}.", "ROMAN COMMENT": ""}),
        dict(ID="X.SHOE.CTA", Trigger="Per shoe card (always)",
             **{"Sentence template": "Check details and availability",
                "Concrete example": "(button at bottom of every shoe card, links to /shoe/{slug})",
                "Mentions shoe?": "Implicit", "Origin": "Live ScanResult.jsx", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="X.HEADER", Trigger="Page header (always)",
             **{"Sentence template": "climbing-gear.com / Your Foot Profile / Scan analysis: sole & side view",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED",
                "Notes": "Em dash removed from subtitle.", "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── Build ───────────────────────────────────────────────────────
def main():
    wb = Workbook()
    build_cover(wb)
    build_section1(wb)
    build_section2(wb)
    build_section3(wb)
    build_card_p1(wb)
    build_card_p2(wb)
    build_card_p3(wb)
    build_cross(wb)
    wb.save(OUT)
    print(f"# wrote {OUT} ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
