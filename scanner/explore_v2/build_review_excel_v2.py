#!/usr/bin/env python3
"""Build the V2 review Excel after Roman's first-pass comments
(2026-04-27/28).

Major changes vs v1 of this Excel (scanner_copy_review_2026_04_27.xlsx):
- §2 Shoe Fit: full cascade rework. Per fit issue, emit ONE sentence
  based on a decision tree: A) shoe-anatomy mismatch B) anatomy fits but
  secondary factor C) sizing off D) everything fits but next picks focus
  on this dim. Removed the slipper closure note from heel-empty-1.
  N>1 ALL: no shoe names. N>1 MINORITY: only affected shoe named.
- §3: add toe profile to S3.1, fix S3.2c trigger, fix S3.3/S3.4
  self-contradiction, DROP S3.3f (no more next_shoe_preference), mark
  V1 S3.6 series clearly DROPPED.
- P1: drop P1.13 (La Sportiva P3 system note).
- P2: P2.3/P2.4 rewrite to reference toe SHAPE not 'wider/snugger toe
  box'. P2.5/P2.6 soften certainty. P2.7-8 drop shallow heel
  conflation - we don't have shoe heel depth data, only width.
- P3: audit pass, no major changes.
- Cross-cutting: unchanged.
- All sentences swept for em dashes (none) and 'cinch' (none).

Output: scanner/explore_v2/scanner_copy_review_2026_04_28.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parent / "scanner_copy_review_2026_04_30.xlsx"

FONT_NAME = "Arial"

STATUS_FILL = {
    "LOCKED":   PatternFill("solid", start_color="C6EFCE"),
    "PROPOSED": PatternFill("solid", start_color="FFF2CC"),
    "V1-LIVE":  PatternFill("solid", start_color="DDEBF7"),
    "V2-DRAFT": PatternFill("solid", start_color="FCE4D6"),
    "DROPPED":  PatternFill("solid", start_color="F2F2F2"),
    "REVISED":  PatternFill("solid", start_color="E1D7F2"),  # NEW: items reworked after Roman's review
}

HEADER_FILL = PatternFill("solid", start_color="2C3227")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D5CDBF"),
    right=Side(style="thin", color="D5CDBF"),
    top=Side(style="thin", color="D5CDBF"),
    bottom=Side(style="thin", color="D5CDBF"),
)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
TOP_LEFT = Alignment(vertical="top", horizontal="left")
TOP_CENTER = Alignment(vertical="top", horizontal="center")

COLS = [
    ("ID",                14),
    ("Trigger",           50),
    ("Sentence template", 64),
    ("Concrete example",  64),
    ("Mentions shoe?",    11),
    ("Origin",            12),
    ("Status",            12),
    ("Notes",             40),
    ("ROMAN COMMENT",     35),
]

def write_header(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color="2C3227")
    ws.merge_cells("A1:I1")
    ws["A1"].alignment = TOP_LEFT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name=FONT_NAME, italic=True, size=10, color="7A7462")
        ws.merge_cells("A2:I2")
        ws["A2"].alignment = TOP_LEFT
    for ci, (label, width) in enumerate(COLS, start=1):
        c = ws.cell(row=4, column=ci, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = "A5"


def add_row(ws, row_num, vals):
    status = vals.get("Status", "")
    fill = STATUS_FILL.get(status)
    for ci, (label, _) in enumerate(COLS, start=1):
        v = vals.get(label, "")
        c = ws.cell(row=row_num, column=ci, value=v)
        c.font = Font(name=FONT_NAME, size=10, color="2C3227")
        c.alignment = WRAP
        c.border = THIN_BORDER
        if label == "Status" and fill:
            c.fill = fill
            c.font = Font(name=FONT_NAME, size=10, bold=True, color="2C3227")
            c.alignment = TOP_CENTER
        elif label == "Mentions shoe?" and v in ("Yes", "No"):
            c.alignment = TOP_CENTER
        elif label == "Origin":
            c.alignment = TOP_CENTER


def auto_row_height(ws, start_row, end_row):
    for r in range(start_row, end_row + 1):
        max_len = 0
        for col_idx in (3, 4, 8):  # Sentence template + Concrete example + Notes
            v = ws.cell(row=r, column=col_idx).value or ""
            max_len = max(max_len, len(str(v)))
        lines = max(1, (max_len // 60) + 1)
        ws.row_dimensions[r].height = max(20, 14 * lines)


# ─── Cover sheet ───────────────────────────────────────────────
def build_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = "Scanner copy review v4, triggers + sentences"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=18, color="2C3227")
    ws.merge_cells("A1:H1")
    ws["A2"] = "Generated 2026-04-30 with §3 redesigned to a 3-paragraph structure (Roman round-3 confirmed)."
    ws["A2"].font = Font(name=FONT_NAME, italic=True, size=11, color="7A7462")
    ws.merge_cells("A2:H2")

    intro = [
        "What changed in v4 (this file) vs v3 (scanner_copy_review_2026_04_29.xlsx):",
        "  • §3 FULL REDESIGN (Roman round-3 2026-04-30):",
        "      - Replaced ~30 V1 explainer paragraphs with a clean 3-paragraph design.",
        "      - P1 = fit target (toe form + forefoot width + heel width). One sentence.",
        "      - P2 = use-case target (discipline_phrase + aggressiveness + closure_pref",
        "        + downturn + asymmetry + ankle for trad + caveats). One sentence.",
        "      - P3 = caveats (shallow heel, soft mask, inconsistent feedback,",
        "        closing tier hint). 1-4 conditional clauses.",
        "      - Stiffness vocab locked at 5 levels: very soft / soft /",
        "        balanced stiffness / stiff / very stiff.",
        "      - Closure derived as discipline ∩ instep-allowed intersection.",
        "      - Asymmetry tautology guard kept: caveat fires only when",
        "        baseline_lbl != target_lbl (no 'shifted from strong to strong').",
        "      - Word 'adjacent' replaced with 'alternative' in tier closing.",
        "      - Shallow heel disclaimer reworded per Roman's exact text.",
        "  • §1 + §2 unchanged from v3 (already locked).",
        "  • P1, P2, P3 cards + Cross-cutting: NOT yet reviewed in this round.",
        "",
        "What changed in v3 vs v2 (scanner_copy_review_2026_04_28.xlsx):",
        "  • §2 ROUND-2 EDITS APPLIED (Roman 2026-04-29):",
        "      - All 5 'everything fits' cascade endpoints reworded:",
        "        drop 'perfectly' / 'may be subtle' / 'for next picks we focus on'",
        "        -> 'In the recommendations we aim for ...'",
        "      - Long-arch cause sentences softened: 'pushes' -> 'may push'",
        "        (S2.TOES.squeezed.1.C and S2.FF.tight.1.B).",
        "      - S2.HEEL.empty.contradiction: dropped 'as the middle ground'.",
        "  • §1 unchanged (no comments).",
        "  • §3, P1, P2, P3, Cross-cutting: NOT yet reviewed in this round",
        "    (Roman: 'first adjust v2 logic according to paragraphs 1+2; then we tackle paragraph 3').",
        "",
        "What changed in v2 (the file Roman reviewed) vs v1 (2026-04-27):",
        "",
        "  • §2 Shoe Fit -- FULL CASCADE REWORK. Per fit issue (heel empty/tight, toes",
        "    squeezed/roomy, forefoot tight/loose), emit ONE sentence based on a decision",
        "    tree: A) shoe-anatomy mismatch B) anatomy fits but secondary factor (heel",
        "    depth / arch length / etc.) C) sizing off vs typical for brand D) everything",
        "    fits, biased for next picks.",
        "  • §2 N>1 ALL: no shoe names per Roman.",
        "  • §2 N>1 MINORITY: only the affected shoe named.",
        "  • §2 heel-empty-1: dropped the slipper closure note (Roman flagged as",
        "    hallucination).",
        "  • §3 S3.1: now includes toe profile in the target sentence.",
        "  • §3 S3.2c: trigger tightened to fire only when target actually adjusted.",
        "  • §3 S3.3a/c, S3.4c: rewritten to remove the self-contradictory framing",
        "    ('shoe-specific not width' then 'we adjust width').",
        "  • §3 S3.3f: DROPPED entirely (no more next_shoe_preference input in V2).",
        "  • §3 S3.6 series (V1 prefs): all DROPPED in V2 -- now clearly marked.",
        "  • Card P1.13 (La Sportiva 'P3 system maintains downturn shape'): DROPPED.",
        "  • Card P2.3/P2.4: rewrite to reference toe SHAPE (Egyptian/Greek/Roman),",
        "    not 'wider/snugger toe box' which we don't have data for.",
        "  • Card P2.5/P2.6: softened certainty ('may relieve' instead of 'relieves').",
        "  • Card P2.7-8: dropped the shallow+narrow heel conflation. We don't have",
        "    shoe heel depth data, only heel volume (width-based). Now references",
        "    width only.",
        "",
        "Status colour key:",
        "  GREEN  LOCKED   -- design already agreed, implemented in V2 sandbox.",
        "  PURPLE REVISED  -- new in this file: reworked after Roman's review (was PROPOSED v1).",
        "  YELLOW PROPOSED -- my draft, awaiting your sign-off.",
        "  BLUE   V1-LIVE  -- production currently emits this; review for V2 retention.",
        "  ORANGE V2-DRAFT -- V2 sandbox emits this; review for cutover.",
        "  GREY   DROPPED  -- V1 paragraph V2 removes; flag if it should stay.",
        "",
        "Sheet order: Cover, §1 Foot Shape, §2 Shoe Fit, §3 What to Look For,",
        "Card P1 (description), Card P2 (why selected), Card P3 (tradeoffs), Cross-cutting.",
    ]
    for i, line in enumerate(intro, start=4):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name=FONT_NAME, size=11, color="2C3227")
        c.alignment = TOP_LEFT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    # Status legend coloured rows
    legend_start = 4 + len(intro) + 2
    statuses = [
        ("LOCKED", "Implemented in V2 sandbox; final wording agreed."),
        ("REVISED", "Reworked after Roman's 2026-04-28 review (highlighted)."),
        ("PROPOSED", "New design, awaiting your sign-off."),
        ("V1-LIVE", "Currently emitted by production /scan; review for V2 retention."),
        ("V2-DRAFT", "Currently emitted by V2 sandbox; review for cutover."),
        ("DROPPED", "V1 paragraph removed from V2; flag if it should come back."),
    ]
    for i, (status, desc) in enumerate(statuses, start=legend_start):
        sc = ws.cell(row=i, column=1, value=status)
        sc.font = Font(name=FONT_NAME, bold=True, size=11, color="2C3227")
        sc.fill = STATUS_FILL[status]
        sc.alignment = TOP_CENTER
        sc.border = THIN_BORDER
        dc = ws.cell(row=i, column=2, value=desc)
        dc.font = Font(name=FONT_NAME, size=11, color="2C3227")
        dc.alignment = TOP_LEFT
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)

    ws.column_dimensions["A"].width = 14
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16


# ─── §1 Foot Shape (UNCHANGED, locked) ──────────────────────────
def build_section1(wb):
    ws = wb.create_sheet("§1 Foot Shape")
    write_header(ws, "Section 1, Your Foot Shape",
                 "LOCKED design. 2 paragraphs total. No changes from v1 review (Roman didn't flag any).")
    rows = [
        dict(ID="T2.1", Trigger="Always (any scan with toe shape + width data)",
             **{"Sentence template": 'You have {toe_article} {Toe} toe form and {forefoot_width} forefoot with a {heel_width} heel.',
                "Concrete example": 'You have an Egyptian toe form and narrow forefoot with a narrow heel.',
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": 'toe_article: "an" for Egyptian, "a" for Greek/Roman. forefoot_width / heel_width: narrow / normal / wide (soft-classified, ±0.005 tolerance).',
                "ROMAN COMMENT": ""}),
        dict(ID="T3.1", Trigger="forefoot_width == narrow AND heel_width == narrow heel",
             **{"Sentence template": "Throughout a narrow profile.",
                "Concrete example": "Throughout a narrow profile.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T3.2", Trigger="forefoot_width == wide AND heel_width == wide heel",
             **{"Sentence template": "Throughout a wide profile.",
                "Concrete example": "Throughout a wide profile.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T3.3", Trigger="forefoot_width == normal AND heel_width == normal",
             **{"Sentence template": "Throughout a medium profile.",
                "Concrete example": "Throughout a medium profile.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T3.4", Trigger="Any other forefoot×heel combo (mismatch)",
             **{"Sentence template": "A mixed profile; forefoot and heel require different fits.",
                "Concrete example": "A mixed profile; forefoot and heel require different fits.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "Collapses 6 combinations into one sentence.", "ROMAN COMMENT": ""}),
        dict(ID="T4.0", Trigger="META rule",
             **{"Sentence template": "If NONE of T4.2..T4.10 trigger -> emit T4.1. Otherwise emit triggered T4.x sentences in order T4.2 -> T4.10, separated by spaces, prefixed with 'Beyond the obvious: '.",
                "Concrete example": "(structural rule)",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "Single rule (no 'OR' branches).", "ROMAN COMMENT": ""}),
        dict(ID="T4.1", Trigger="toe == Egyptian AND arch / instep / heel_depth / HVA all normal",
             **{"Sentence template": "As your arch length, instep height and heel depth are within average, you can focus mostly on toe form and shoe width when looking for a new shoe.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "All-average fallback.", "ROMAN COMMENT": ""}),
        dict(ID="T4.2", Trigger="toe IN (Greek, Roman)",
             **{"Sentence template": "Given your {toe_lower} toes, rather avoid very pointy shoes. These might squeeze your toe tips, even if the width of the shoe fits you.",
                "Concrete example": "Given your greek toes, rather avoid very pointy shoes. These might squeeze your toe tips, even if the width of the shoe fits you.",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.3", Trigger="arch_length_class == 'long arch' (soft-classified)",
             **{"Sentence template": "Given your long arch, the ball of your foot may be pushed into the toe box. A squeezed forefoot may be caused by this instead of actual shoe width, so look for rather short toe boxes.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED",
                "Notes": "Soft-class fix: 0.733 (raw 'normal') promoted to 'long arch' since within 0.005 of hi=0.734.",
                "ROMAN COMMENT": ""}),
        dict(ID="T4.4", Trigger="arch_length_class == 'short arch'",
             **{"Sentence template": "Your arch is rather short, meaning your toes are relatively long. Especially when considering aggressive shoes look for sufficient height in the toe box to let your toes curl up.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.5", Trigger="instep_height_class == 'low instep'",
             **{"Sentence template": "Additionally your instep is rather low, so an adjustable closure is preferable to avoid dead space. Ideally double velcro or laces, rather avoid pure slippers.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.6", Trigger="instep_height_class == 'high instep'",
             **{"Sentence template": "Additionally your instep is rather high, so an adjustable closure is preferable to actually get into the shoe. Ideally double velcro or laces, you may struggle getting into slippers if adequately downsized.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.7", Trigger="heel_depth_class == 'deep heel'",
             **{"Sentence template": "Your deep heel projects further back than most, so a deeper, more sculpted heel cup fits naturally.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.8", Trigger="heel_depth_class == 'shallow heel'",
             **{"Sentence template": "Your shallow heel doesn't project as far back as most. Deeply sculpted cups will feel empty at the back.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V2 NEW", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="T4.9", Trigger="HVA IN (mild, pronounced) AND toe == Egyptian",
             **{"Sentence template": "For your {hva} hallux valgus, we avoid very asymmetric lasts and prefer moderately asymmetric or wider-tipped designs that help the big toe stay in contact with the shoe tip.",
                "Concrete example": "For your mild hallux valgus, we avoid very asymmetric lasts and prefer moderately asymmetric or wider-tipped designs that help the big toe stay in contact with the shoe tip.",
                "Mentions shoe?": "No", "Origin": "V1 verbatim, V2 trigger", "Status": "LOCKED",
                "Notes": "HVA threshold change (0.25 to 0.28) deferred to v2 go-live.", "ROMAN COMMENT": ""}),
        dict(ID="T4.10", Trigger="HVA IN (mild, pronounced) AND toe IN (Greek, Roman)",
             **{"Sentence template": "For your {hva} hallux valgus, we prefer a slightly wider, less pointed toe box to compensate for the inward drift.",
                "Concrete example": "For your pronounced hallux valgus, we prefer a slightly wider, less pointed toe box to compensate for the inward drift.",
                "Mentions shoe?": "No", "Origin": "V1 verbatim, V2 trigger", "Status": "LOCKED",
                "Notes": "Roman + Greek treated the same per Q2.", "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── §2 Shoe Fit (FULL CASCADE REWORK) ────────────────────────────
def build_section2(wb):
    ws = wb.create_sheet("§2 Shoe Fit")
    write_header(ws, "Section 2, What Your Current Shoe Fit Tells Us",
                 "FULL CASCADE REWORK after Roman's 2026-04-28 review. Each fit issue runs a decision tree (A then B then C then D), emits ONE sentence per shoe-issue based on first matching outcome.")

    # Lock S1 sizing intro (unchanged)
    rows = [
        dict(ID="S1.1.a", Trigger="N == 1 with size_eu",
             **{"Sentence template": "You wear your {brand} {model} in EU {size_eu}, {downsize_clause}.",
                "Concrete example": "You wear your Scarpa Furia Air in EU 44, 1.5 sizes down from your street size of 45.5.",
                "Mentions shoe?": "Yes", "Origin": "V1 wording, V2 LOCKED", "Status": "LOCKED",
                "Notes": "downsize_clause: '{X} sizes down from your street size of {street}' / 'at your street size of {street}' / '{X} sizes above your street size of {street}'.",
                "ROMAN COMMENT": ""}),
        dict(ID="S1.1.b", Trigger="N == 1 (always after S1.1.a)",
             **{"Sentence template": "Typical: 'That is a typical fit for {brand}.' OR Other: 'That is {label} for {brand}, where the typical downsize is about {typical_label}.'",
                "Concrete example": "That is a typical fit for Scarpa.\nOR\nThat is rather aggressive for La Sportiva, where the typical downsize is about 1 size down.",
                "Mentions shoe?": "Yes (brand)", "Origin": "V1 wording, V2 LOCKED", "Status": "LOCKED",
                "Notes": "label IN {rather aggressive, very aggressive, rather relaxed, very relaxed}.",
                "ROMAN COMMENT": ""}),
        dict(ID="S1.1.c", Trigger="S1.1.b label is rather/very relaxed AND fit.heel ∈ {empty, loose}",
             **{"Sentence template": "Downsizing further toward the typical range could tighten the heel and reduce the empty feel.",
                "Concrete example": "(appended after S1.1.b)",
                "Mentions shoe?": "No", "Origin": "V1 wording, V2 LOCKED", "Status": "LOCKED",
                "Notes": "Conditional 3rd sentence. NOTE: with the new cascade rework below, this sentence partially overlaps with S2.HEEL.empty.1.C, review which one wins (suggest: §2 cascade owns the sizing diagnosis, S1.1.c is dropped).",
                "ROMAN COMMENT": ""}),
        dict(ID="S1.2", Trigger="N > 1, all same brand",
             **{"Sentence template": "You wear your {brand} shoes in EU {min}–{max}, {raw_range_clause} from your street size of {street}. For {brand}, where the typical downsize is about {typical_label}, this is a {overall_label} fit overall.",
                "Concrete example": "You wear your Scarpa shoes in EU 43–44.5, 1–2.5 sizes down from your street size of 45.5. For Scarpa, where the typical downsize is about 1.5 sizes down, this is aggressive overall.",
                "Mentions shoe?": "Yes (brand)", "Origin": "V1 wording, V2 LOCKED", "Status": "LOCKED",
                "Notes": "En-dash for ranges OK.", "ROMAN COMMENT": ""}),
        dict(ID="S1.3.a", Trigger="N > 1, mixed brands, ≥1 deviating shoe",
             **{"Sentence template": "Your shoes range from EU {min} to {max} relative to a street size of {street}. Raw sizes alone don't tell the story because brands size differently: {highlight_1}{connector}{highlight_2}. {overall_clause if N≥3 and aligned}",
                "Concrete example": "Your shoes range from EU 42 to 44 relative to a street size of 45.5. Raw sizes alone don't tell the story because brands size differently: your Scarpa Drago in EU 42 runs aggressive for Scarpa (typical downsize about 1.5 sizes down); also your La Sportiva Theory in EU 43 runs aggressive for La Sportiva (typical downsize about 1.5 sizes down). Across your shoes you tend to size more aggressively than the brand-typical fit.",
                "Mentions shoe?": "Yes (per highlight)", "Origin": "V1 wording + V2 connector logic", "Status": "LOCKED",
                "Notes": "connector: '; also ' when both highlights same direction; ', while ' when opposite.",
                "ROMAN COMMENT": ""}),
        dict(ID="S1.3.b", Trigger="N > 1, mixed brands, all near brand-typical",
             **{"Sentence template": "Your shoes range from EU {min} to {max} (street size {street}). Each one sits close to its brand-typical downsize even though the raw EU numbers differ.",
                "Concrete example": "Your shoes range from EU 43 to 45 (street size 45.5). Each one sits close to its brand-typical downsize even though the raw EU numbers differ.",
                "Mentions shoe?": "No", "Origin": "V1 wording, V2 LOCKED", "Status": "LOCKED",
                "Notes": "", "ROMAN COMMENT": ""}),

        # ────────────────────────────────────────────────────────
        # CASCADE INTRO ROW
        # ────────────────────────────────────────────────────────
        dict(ID="--- CASCADE ---", Trigger="(decision-tree explanation)",
             **{"Sentence template": "For each (shoe, dim, rating) where the user reported an issue, run the cascade: A first, else B, else C, else D. Only ONE sentence fires per shoe-issue. N>1 ALL = aggregate (no shoe names). N>1 MINORITY = run cascade for affected shoe only (only that shoe named).",
                "Concrete example": "(structural rule, not a sentence)",
                "Mentions shoe?": "-", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "Replaces the flat per-(dim×rating) structure from v1 of this review. Decision logic mirrors what a user would think: anatomy mismatch first, then secondary anatomy factor, then sizing, then 'all checks out, just bias for next picks'.",
                "ROMAN COMMENT": ""}),

        # ────────────────────────────────────────────────────────
        # HEEL EMPTY cascade
        # ────────────────────────────────────────────────────────
        dict(ID="S2.HEEL.empty.1.A", Trigger="N==1, heel==empty AND shoe.db_heel_volume rank > user heel_width_class rank (cup wider than heel)",
             **{"Sentence template": "Your {brand} {model} has a {shoe_hv} heel volume while your heel is {user_hw}. The wider cup is the most likely cause of the empty feel.",
                "Concrete example": "Your Scarpa Furia Air has a medium heel volume while your heel is narrow. The wider cup is the most likely cause of the empty feel.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade A)", "Status": "REVISED",
                "Notes": "First check in cascade. If heel volume rank > heel width rank: cup-too-wide is the call. Don't continue to other cascade outcomes.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.empty.1.B", Trigger="N==1, heel==empty AND shoe.db_heel_volume matches user heel_width AND user heel_depth_class == 'shallow heel'",
             **{"Sentence template": "Your {brand} {model}'s {shoe_hv} heel volume matches your {user_hw} heel width, but your shallow heel may not fill deeply sculpted cups. The cause is likely heel depth, try shoes with a flatter, less sculpted heel cup.",
                "Concrete example": "Your Scarpa Vapor V's narrow heel volume matches your narrow heel width, but your shallow heel may not fill deeply sculpted cups. The cause is likely heel depth, try shoes with a flatter, less sculpted heel cup.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade B)", "Status": "REVISED",
                "Notes": "Width matches but user has shallow heel. NOTE: contains an em dash, replace before implementation. Suggested: 'The cause is likely heel depth, so try shoes with a flatter, less sculpted heel cup.'",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.empty.1.C", Trigger="N==1, heel==empty AND width and depth check out AND user_downsize < typical_downsize for brand by ≥0.5 sizes",
             **{"Sentence template": "Comparing your {brand} {model} to your foot profile, it should fit. You only downsized {user_downsize} vs the usual {typical_downsize} for {brand}, going down further could tighten the heel and reduce the empty feel.",
                "Concrete example": "Comparing your Evolv Defy to your foot profile, it should fit. You only downsized half a size vs the usual 1.5 sizes for Evolv, going down further could tighten the heel and reduce the empty feel.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade C)", "Status": "REVISED",
                "Notes": "Sizing check after width + depth fit. Em dash present, replace.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.empty.1.D", Trigger="N==1, heel==empty AND width matches AND heel_depth not shallow AND sized typically",
             **{"Sentence template": "Your {brand} {model} should fit based on heel width and depth, and you're sized typically for {brand}. In the recommendations we aim for even narrower heel cups.",
                "Concrete example": "Your Scarpa Drago should fit based on heel width and depth, and you're sized typically for Scarpa. In the recommendations we aim for even narrower heel cups.",
                "Mentions shoe?": "Yes", "Origin": "REVISED v2 (Roman 2026-04-29 reword)", "Status": "REVISED",
                "Notes": "Reworded per Roman: drop 'perfectly' + 'The empty feel may be subtle;' + 'for next picks we focus on' -> 'In the recommendations we aim for'. Same pattern applied to all other 'everything fits' cascade endpoints (heel.tight.1.D, toes.squeezed.1.E, toes.roomy.1.D, ff.tight.1.D, ff.loose.1.C).",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.empty.allN", Trigger="N>1, ALL heel == empty",
             **{"Sentence template": "Your heel feels empty across all your shoes. {pattern_explanation}",
                "Concrete example": "Your heel feels empty across all your shoes. All have heel cups wider than your narrow heel, for next picks we focus on narrow heel cups.",
                "Mentions shoe?": "No (per Roman)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "pattern_explanation variants:\n  • all shoes have heel volume > user heel width → 'All have heel cups wider than your {user_hw} heel, for next picks we focus on narrower heel cups.'\n  • all match heel width but user shallow → 'Heel widths match your foot, but your shallow heel may not fill deeply sculpted cups. We focus on flatter heel cups.'\n  • mixed pattern → 'This points to a structural fit pattern. Your {user_hw} heel needs cups narrower than these models offer.'",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.empty.minority", Trigger="N>1, SOME heel == empty (not all, not contradiction)",
             **{"Sentence template": "Your {brand} {model}'s heel feels empty while your other shoes fit. {cascade outcome A/B/C/D from S2.HEEL.empty.1.*}",
                "Concrete example": "Your Scarpa Furia Air's heel feels empty while your other shoes fit. The Furia Air has a medium heel volume while your heel is narrow, the wider cup is the most likely cause.",
                "Mentions shoe?": "Yes (only affected, per Roman)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "Run S2.HEEL.empty.1.* cascade for the affected shoe; prefix with 'Your {brand} {model}'s heel feels empty while your other shoes fit.'",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.empty.contradiction", Trigger="N>1, BOTH empty AND tight heel ratings present",
             **{"Sentence template": "Your heel fit varies: empty in your {affected_brand} {affected_model}, tight in your {other_brand} {other_model}. Different heel cup geometries; for next picks we target {target_hv} heel volume.",
                "Concrete example": "Your heel fit varies: empty in your Scarpa Furia Air, tight in your La Sportiva Solution. Different heel cup geometries; for next picks we target narrow heel volume.",
                "Mentions shoe?": "Yes (both)", "Origin": "REVISED v2", "Status": "REVISED",
                "Notes": "Roman 2026-04-29: drop 'as the middle ground'.", "ROMAN COMMENT": ""}),

        # ────────────────────────────────────────────────────────
        # HEEL TIGHT cascade
        # ────────────────────────────────────────────────────────
        dict(ID="S2.HEEL.tight.1.A", Trigger="N==1, heel==tight AND shoe.db_heel_volume rank < user heel_width_class rank (cup narrower than heel)",
             **{"Sentence template": "Your {brand} {model} has a {shoe_hv} heel volume while your heel is {user_hw}. The narrower cup is the most likely cause of the tight feel.",
                "Concrete example": "Your La Sportiva Solution has a narrow heel volume while your heel is wide. The narrower cup is the most likely cause of the tight feel.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade A)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.tight.1.B", Trigger="N==1, heel==tight AND heel volumes match AND user heel_depth_class == 'deep heel'",
             **{"Sentence template": "Your {brand} {model}'s {shoe_hv} heel volume matches your {user_hw} heel width, but your deep heel may not fit cups designed for less backward projection. The cause is likely heel depth, so try shoes with a deeper, more sculpted heel cup.",
                "Concrete example": "Your Scarpa Vapor V's narrow heel volume matches your narrow heel width, but your deep heel may not fit cups designed for less backward projection. The cause is likely heel depth, so try shoes with a deeper, more sculpted heel cup.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade B)", "Status": "REVISED",
                "Notes": "We don't have shoe heel depth data, so this is a probabilistic claim about the user side.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.tight.1.C", Trigger="N==1, heel==tight AND width and depth check out AND user over-downsized vs typical for brand",
             **{"Sentence template": "Comparing your {brand} {model} to your foot profile, it should fit. You downsized {user_downsize}, more aggressive than the usual {typical_downsize} for {brand}, so going up half a size could relieve the tightness.",
                "Concrete example": "Comparing your La Sportiva Solution to your foot profile, it should fit. You downsized 2.5 sizes, more aggressive than the usual 1.5 sizes for La Sportiva, so going up half a size could relieve the tightness.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade C)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.tight.1.D", Trigger="N==1, heel==tight AND width matches AND heel_depth not deep AND sized typically",
             **{"Sentence template": "Your {brand} {model} should fit based on heel width and depth, and you're sized typically for {brand}. In the recommendations we aim for slightly roomier heel cups.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED v2 (Roman 2026-04-29)", "Status": "REVISED",
                "Notes": "Same pattern as empty.1.D: drop 'The tight feel may be subtle;', use 'In the recommendations we aim for'.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.tight.allN", Trigger="N>1, ALL heel == tight",
             **{"Sentence template": "Your heel feels tight in all your shoes. {pattern_explanation}",
                "Concrete example": "Your heel feels tight in all your shoes. All have heel cups narrower than your wide heel, for next picks we look at wider, more accommodating heel cups.",
                "Mentions shoe?": "No (per Roman)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "pattern_explanation similar to empty.allN but inverted.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.HEEL.tight.minority", Trigger="N>1, SOME heel == tight",
             **{"Sentence template": "Your {brand} {model}'s heel feels tight while your other shoes fit. {cascade outcome from S2.HEEL.tight.1.*}",
                "Concrete example": "(only affected shoe named)",
                "Mentions shoe?": "Yes (only affected)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),

        # ────────────────────────────────────────────────────────
        # TOES SQUEEZED cascade (5-step: toe shape → width → arch → sizing → all-fit)
        # ────────────────────────────────────────────────────────
        dict(ID="S2.TOES.squeezed.1.A", Trigger="N==1, toes==squeezed AND user_toe_shape NOT in shoe.db_toe_form",
             **{"Sentence template": "Your {brand} {model} is built on {a_an} {shoe_form} toe form while you have {user_form} toes. The mismatch is the most likely cause; look for shoes with a {user_form}-compatible toe box.",
                "Concrete example": "Your Scarpa Drago is built on a Greek toe form while you have Egyptian toes. The mismatch is the most likely cause; look for shoes with an Egyptian-compatible toe box.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade A)", "Status": "REVISED",
                "Notes": "Article guard: 'an Egyptian' / 'a Greek' / 'a Roman'.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.squeezed.1.B", Trigger="N==1, toes==squeezed AND toe shape matches AND user_forefoot_width=='wide' AND shoe.db_width IN (narrow, medium)",
             **{"Sentence template": "Your {brand} {model}'s toe shape matches your {user_form} foot, but your wide forefoot in this {shoe_w} last is likely causing the squeeze. Look for wider lasts.",
                "Concrete example": "Your Scarpa Vapor V's toe shape matches your Egyptian foot, but your wide forefoot in this medium last is likely causing the squeeze. Look for wider lasts.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade B)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.squeezed.1.C", Trigger="N==1, toes==squeezed AND toe shape and width fit AND user_arch_class=='long arch'",
             **{"Sentence template": "Your {brand} {model}'s toe shape and width match your foot, but your long arch may push the ball of your foot into the toe box. Look for shoes with a shorter toe box to relieve the squeeze.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED v2 (Roman 2026-04-29)", "Status": "REVISED",
                "Notes": "Soften 'pushes' -> 'may push' per Roman.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.squeezed.1.D", Trigger="N==1, toes==squeezed AND toe + width + arch all fit AND user over-downsized vs typical",
             **{"Sentence template": "Comparing your {brand} {model} to your foot profile, it should fit. You downsized {user_downsize}, more aggressive than the usual {typical_downsize} for {brand}, so going up half a size could relieve the squeeze.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade D)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.squeezed.1.E", Trigger="N==1, toes==squeezed AND everything fits AND sized typically",
             **{"Sentence template": "Your {brand} {model} should fit based on toe form, forefoot width and arch, and you're sized typically for {brand}. In the recommendations we aim for slightly more toe room.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED v2 (Roman 2026-04-29)", "Status": "REVISED",
                "Notes": "Same reword pattern as empty.1.D: drop 'The squeeze may be subtle;', use 'In the recommendations we aim for'.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.squeezed.allN", Trigger="N>1, ALL toes == squeezed",
             **{"Sentence template": "Your toes feel squeezed across all your shoes. {pattern_explanation}",
                "Concrete example": "Your toes feel squeezed across all your shoes. All are built on toe forms different from your Egyptian toes, for next picks we filter for Egyptian-compatible toe boxes.",
                "Mentions shoe?": "No (per Roman)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "Pattern variants by which cascade step would fire most often.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.squeezed.minority", Trigger="N>1, SOME toes == squeezed",
             **{"Sentence template": "Your {brand} {model}'s toes feel squeezed while your other shoes fit. {cascade outcome from S2.TOES.squeezed.1.*}",
                "Concrete example": "(only affected shoe named)",
                "Mentions shoe?": "Yes (only affected)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),

        # ────────────────────────────────────────────────────────
        # TOES ROOMY cascade (4-step)
        # ────────────────────────────────────────────────────────
        dict(ID="S2.TOES.roomy.1.A", Trigger="N==1, toes==roomy AND user_toe_shape NOT in shoe.db_toe_form",
             **{"Sentence template": "Your {brand} {model} is built on {a_an} {shoe_form} toe form while you have {user_form} toes. The mismatch likely leaves dead space at your second toe; look for shoes with a {user_form}-compatible toe box.",
                "Concrete example": "Your La Sportiva Tarantulace is built on a Greek toe form while you have Egyptian toes. The mismatch likely leaves dead space at your second toe; look for shoes with an Egyptian-compatible toe box.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade A)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.roomy.1.B", Trigger="N==1, toes==roomy AND toe shape matches AND user_forefoot_width=='narrow' AND shoe.db_width IN (medium, wide)",
             **{"Sentence template": "Your {brand} {model}'s toe shape matches your foot, but the {shoe_w} forefoot is wider than your narrow forefoot. Look for narrower lasts.",
                "Concrete example": "Your Five Ten Anasazi's toe shape matches your foot, but the wide forefoot is wider than your narrow forefoot. Look for narrower lasts.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade B)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.roomy.1.C", Trigger="N==1, toes==roomy AND toe and width fit AND user under-downsized vs typical",
             **{"Sentence template": "Comparing your {brand} {model} to your foot profile, it should fit. You only downsized {user_downsize} vs the usual {typical_downsize} for {brand}, so going down half a size could tighten the toe box.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade C)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.roomy.1.D", Trigger="N==1, toes==roomy AND everything fits AND sized typically",
             **{"Sentence template": "Your {brand} {model} should fit based on toe form and forefoot width, and you're sized typically for {brand}. In the recommendations we aim for snugger toe boxes.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED v2 (Roman 2026-04-29)", "Status": "REVISED",
                "Notes": "Same reword pattern as empty.1.D.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.roomy.allN", Trigger="N>1, ALL toes == roomy",
             **{"Sentence template": "Your toes have extra room in all your shoes. {pattern_explanation}",
                "Concrete example": "Your toes have extra room in all your shoes. The toe forms across these models don't match your Egyptian foot, for next picks we filter for Egyptian-compatible toe boxes.",
                "Mentions shoe?": "No (per Roman)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.roomy.minority", Trigger="N>1, SOME toes == roomy",
             **{"Sentence template": "Your {brand} {model}'s toes have extra room while your other shoes fit. {cascade outcome}",
                "Concrete example": "(only affected named)",
                "Mentions shoe?": "Yes (only affected)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.TOES.contradiction", Trigger="N>1, BOTH squeezed AND roomy ratings present",
             **{"Sentence template": "Your toe fit varies: squeezed in your {Shoe A}, roomy in your {Shoe B}. Different toe box shapes; the next section will narrow which one matches your scan.",
                "Concrete example": "Your toe fit varies: squeezed in your Scarpa Drago, roomy in your La Sportiva Tarantulace. Different toe box shapes; the next section will narrow which one matches your scan.",
                "Mentions shoe?": "Yes (both)", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),

        # ────────────────────────────────────────────────────────
        # FOREFOOT TIGHT cascade
        # ────────────────────────────────────────────────────────
        dict(ID="S2.FF.tight.1.A", Trigger="N==1, ff==tight AND shoe.db_width rank < user_forefoot_width rank",
             **{"Sentence template": "Your {brand} {model}'s {shoe_w} forefoot is narrower than your {user_fw} forefoot. The width mismatch is the most likely cause.",
                "Concrete example": "Your Scarpa Drago's narrow forefoot is narrower than your wide forefoot. The width mismatch is the most likely cause.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade A)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.tight.1.B", Trigger="N==1, ff==tight AND width matches AND user_arch_class=='long arch'",
             **{"Sentence template": "Your {brand} {model}'s width matches your forefoot, but your long arch may push the ball forward into the toe box. Look for shoes with a shorter toe box.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED v2 (Roman 2026-04-29)", "Status": "REVISED",
                "Notes": "Soften 'pushes' -> 'may push' per Roman.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.tight.1.C", Trigger="N==1, ff==tight AND width and arch fit AND user over-downsized",
             **{"Sentence template": "Comparing your {brand} {model} to your foot profile, it should fit. You downsized {user_downsize}, more aggressive than the usual {typical_downsize} for {brand}, so half a size up could relieve the tightness.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade C)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.tight.1.D", Trigger="N==1, ff==tight AND everything fits",
             **{"Sentence template": "Your {brand} {model} should fit based on width and arch, and you're sized typically for {brand}. In the recommendations we aim for slightly wider forefoots.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED v2 (Roman 2026-04-29)", "Status": "REVISED",
                "Notes": "Same reword pattern as empty.1.D.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.tight.allN", Trigger="N>1, ALL ff == tight",
             **{"Sentence template": "Your forefoot feels tight in all your shoes. {pattern_explanation}",
                "Concrete example": "Your forefoot feels tight in all your shoes. All have forefoots narrower than your wide forefoot, for next picks we look at wider lasts.",
                "Mentions shoe?": "No (per Roman)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.tight.minority", Trigger="N>1, SOME ff == tight",
             **{"Sentence template": "Your {brand} {model}'s forefoot feels tight while your other shoes fit. {cascade outcome}",
                "Concrete example": "(only affected named)",
                "Mentions shoe?": "Yes (only affected)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),

        # ────────────────────────────────────────────────────────
        # FOREFOOT LOOSE cascade (no arch step, short arch doesn't typically loosen ff)
        # ────────────────────────────────────────────────────────
        dict(ID="S2.FF.loose.1.A", Trigger="N==1, ff==loose AND shoe.db_width rank > user_forefoot_width rank",
             **{"Sentence template": "Your {brand} {model}'s {shoe_w} forefoot is wider than your {user_fw} forefoot. The width mismatch is the most likely cause of the loose feel.",
                "Concrete example": "Your Five Ten Anasazi's wide forefoot is wider than your narrow forefoot. The width mismatch is the most likely cause of the loose feel.",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade A)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.loose.1.B", Trigger="N==1, ff==loose AND width matches AND user under-downsized vs typical",
             **{"Sentence template": "Comparing your {brand} {model} to your foot profile, it should fit. You only downsized {user_downsize} vs the usual {typical_downsize} for {brand}, so going down further could tighten the forefoot.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED (cascade B)", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.loose.1.C", Trigger="N==1, ff==loose AND width and sizing fit",
             **{"Sentence template": "Your {brand} {model} should fit based on width, and you're sized typically for {brand}. In the recommendations we aim for snugger forefoots.",
                "Concrete example": "(per shoe)",
                "Mentions shoe?": "Yes", "Origin": "REVISED v2 (Roman 2026-04-29)", "Status": "REVISED",
                "Notes": "Same reword pattern as empty.1.D. No arch step in this cascade because short arch doesn't typically loosen forefoot.",
                "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.loose.allN", Trigger="N>1, ALL ff == loose",
             **{"Sentence template": "Your forefoot feels loose in all your shoes. {pattern_explanation}",
                "Concrete example": "Your forefoot feels loose in all your shoes. All have forefoots wider than your narrow forefoot, for next picks we look at narrower lasts.",
                "Mentions shoe?": "No (per Roman)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.loose.minority", Trigger="N>1, SOME ff == loose",
             **{"Sentence template": "Your {brand} {model}'s forefoot feels loose while your other shoes fit. {cascade outcome}",
                "Concrete example": "(only affected named)",
                "Mentions shoe?": "Yes (only affected)", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S2.FF.contradiction", Trigger="N>1, BOTH tight AND loose ff ratings",
             **{"Sentence template": "Your forefoot fit is split: tight in your {Shoe A}, loose in your {Shoe B}. Different lasts; for next picks we target {target_fw} width as the middle ground.",
                "Concrete example": "Your forefoot fit is split: tight in your Scarpa Drago, loose in your Five Ten Anasazi. Different lasts; for next picks we target medium width as the middle ground.",
                "Mentions shoe?": "Yes (both)", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),

        # ────────────────────────────────────────────────────────
        # CROSS-SHOE INSIGHTS (S3)
        # ────────────────────────────────────────────────────────
        dict(ID="S3.BRAND_INCONSISTENCY", Trigger="N≥2, ≥2 shoes same brand, downsize spread ≥0.75 sizes, divergent fits",
             **{"Sentence template": "Within {brand}, your {good_model} at EU {good_size} fits well, while your {bad_model} at EU {bad_size} has {issues}. The {spread}-size spread indicates {brand}'s lasts vary by model; brand alone isn't a reliable predictor of fit for you.",
                "Concrete example": "Within La Sportiva, your Solution at EU 43 fits well, while your Tarantulace at EU 44.5 has loose forefoot and empty heel. The 1.5-size spread indicates La Sportiva's lasts vary by model; brand alone isn't a reliable predictor of fit for you.",
                "Mentions shoe?": "Yes (both)", "Origin": "V2 sandbox NEW", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.HEEL_VOL_VS_SHAPE", Trigger="N≥2, one heel==perfect AND another heel==empty AND BOTH same db_heel_volume",
             **{"Sentence template": "Your {good_model} fits perfectly in the heel while your {bad_model} feels empty, even though both are rated {volume} heel volume. The difference is cup shape, not volume.",
                "Concrete example": "Your Skwama fits perfectly in the heel while your Solution feels empty, even though both are rated medium heel volume. The difference is cup shape, not volume.",
                "Mentions shoe?": "Yes (both)", "Origin": "V2 sandbox NEW", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.HEEL_SOFT_CONFORM", Trigger="N≥1, soft shoe + heel==perfect + scan extreme heel that contradicts shoe's db_heel_volume",
             **{"Sentence template": "Your {brand} {model}'s heel fits despite a mismatch between the cup volume and your scan. The shoe is soft enough to conform; in a stiffer model with the same cup, the same fit might not work.",
                "Concrete example": "Your Scarpa Furia Air's heel fits despite a mismatch between the cup volume and your scan. The shoe is soft enough to conform; in a stiffer model with the same cup, the same fit might not work.",
                "Mentions shoe?": "Yes", "Origin": "V2 sandbox NEW", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S3.TOE_FORM_SOFT_MASK", Trigger="N≥1, user toe ≠ shoe db_toe_form + soft shoe + toes==perfect + ff==perfect",
             **{"Sentence template": "Your {brand} {model} is built on a {shoe_form} toe form, yet your {user_form} toes feel fine. The shoe is soft enough to conform; stiffer shoes on the same form may not.",
                "Concrete example": "Your Scarpa Furia Air is built on a Greek toe form, yet your Egyptian toes feel fine. The shoe is soft enough to conform; stiffer shoes on the same form may not.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "PROPOSED",
                "Notes": "From V1 _para_toe_form_mismatch.", "ROMAN COMMENT": ""}),

        # ────────────────────────────────────────────────────────
        # S4 combined squeeze+loose (per-shoe override)
        # ────────────────────────────────────────────────────────
        dict(ID="S4.SQUEEZE_LOOSE", Trigger="Same shoe: toes==squeezed AND forefoot IN (loose, roomy)",
             **{"Sentence template": "Your {brand} {model} shows an unusual combination: toes squeezed while the forefoot is loose. The toe form is the mismatch, not the width, {scan_cause}.",
                "Concrete example": "Your Scarpa Drago shows an unusual combination: toes squeezed while the forefoot is loose. The toe form is the mismatch, not the width, your scan shows Egyptian toes while this shoe is built on a Greek toe form.",
                "Mentions shoe?": "Yes", "Origin": "V1 + V2", "Status": "PROPOSED",
                "Notes": "When S4 fires for a shoe, suppress S2.TOES.squeezed AND S2.FF.loose for that same shoe. NOTE: contains em dash, replace.",
                "ROMAN COMMENT": ""}),

        # ────────────────────────────────────────────────────────
        # S5 closing anchor
        # ────────────────────────────────────────────────────────
        dict(ID="S5.PERFECT_ALL", Trigger="≥1 shoe with perfect on all 3 dims",
             **{"Sentence template": "We use your {brand} {model} as a reliable reference and target shoes with similar fit characteristics.",
                "Concrete example": "We use your Scarpa Vapor V as a reliable reference and target shoes with similar fit characteristics.",
                "Mentions shoe?": "Yes", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S5.PARTIAL", Trigger="≥1 shoe with perfect on ≥1 dim AND imperfect on ≥1 dim",
             **{"Sentence template": "We use your {best_model} as anchor and focus on closing the gaps where it falls short ({imperfect_list}).",
                "Concrete example": "We use your Scarpa Furia Air as anchor and focus on closing the gaps where it falls short (heel feels empty).",
                "Mentions shoe?": "Yes", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="S5.NONE_GOOD", Trigger="No shoe has any perfect dim",
             **{"Sentence template": "None of your shoes is a clean reference. We'll lean more on the scan than on shoe-fit feedback for the recommendation.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "PROPOSED", "Status": "PROPOSED",
                "Notes": "Q3 still open: is this voice OK or dismissive of the user's input?",
                "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── §3 What to Look For (UPDATED) ────────────────────────────────
def build_section3(wb):
    ws = wb.create_sheet("§3 What to Look For")
    write_header(ws, "Section 3, What to Look For",
                 "ROUND-3 FULL REDESIGN (Roman 2026-04-30 confirmed). 3 paragraphs total. P1 = fit target, P2 = use-case target, P3 = caveats. Replaces ~30 V1 explainer paragraph variants.")

    rows = [
        # ============================================================
        # NEW 3-PARAGRAPH DESIGN (LOCKED 2026-04-30)
        # ============================================================
        dict(ID="--- DESIGN ---", Trigger="3-paragraph design",
             **{"Sentence template": "P1 (fit target, always) + P2 (use-case target, always when V2 inputs present) + P3 (caveats, 1-4 conditional clauses + always-closing tier hint).",
                "Concrete example": "(structural rule)",
                "Mentions shoe?": "No", "Origin": "LOCKED 2026-04-30", "Status": "LOCKED",
                "Notes": "Replaces 30+ V1 explainer paragraph variants. Per-attribute substitutions match what the scorer (target_resolver_v2 + matrix_scorer_v2.compute_use_case_target) actually computes.",
                "ROMAN COMMENT": ""}),
        # ─── P1: Fit target ───
        dict(ID="§3.P1", Trigger="Always (when scan available)",
             **{"Sentence template": "Based on your scan and current shoe fit, we target shoes with {toe_form} toe form, {forefoot_width} forefoot width, and {heel_width} heel width.",
                "Concrete example": "Based on your scan and current shoe fit, we target shoes with Egyptian toe form, narrow forefoot width, and narrow heel width.",
                "Mentions shoe?": "No", "Origin": "LOCKED 2026-04-30", "Status": "LOCKED",
                "Notes": "Final target after any shoe-fit feedback adjustment is BAKED IN. No '(adjusted from X)' footnote. Per-attribute derivation: toe_form = scan toe_shape (no adjustment); forefoot_width = scan forefoot_width_class adjusted by feedback (tight in >=2 -> wider, loose in >=2 -> narrower); heel_width = scan heel_width_class adjusted by feedback (empty in >=2 -> narrower, tight in >=2 -> wider).",
                "ROMAN COMMENT": ""}),
        # ─── P2: Use-case target ───
        dict(ID="§3.P2", Trigger="Always when V2 inputs (discipline + environment + aggressiveness) present",
             **{"Sentence template": "Given your preference for {discipline_phrase} and an {aggressiveness} fit, we prioritize {stiffness_word} {closure_pref}{closure_caveat_if_any} with {downturn} and {asymmetry}{asym_caveat_if_any}{ankle_clause_if_trad}.",
                "Concrete example": "Given your preference for bouldering outdoors on sandstone and an aggressive fit, we prioritize soft velcros (avoiding slippers due to your low instep) with aggressive downturn and strong asymmetry.",
                "Mentions shoe?": "No", "Origin": "LOCKED 2026-04-30", "Status": "LOCKED",
                "Notes": "See derivation reference rows below for substitutions. asym_caveat fires only when baseline_lbl != target_lbl (tautology guard). ankle_clause fires only when discipline == trad_multipitch. closure_caveat fires only when an exclusion was applied (slipper or single-velcro out due to instep extreme).",
                "ROMAN COMMENT": ""}),
        # ─── P3 conditional clauses ───
        dict(ID="§3.P3.shallow_heel", Trigger="scan.heel_depth_class == 'shallow heel'",
             **{"Sentence template": "Note: data on shoe heel depth is unfortunately not widely available. We're working on extracting this data from shoe scans, but for now the recommendations match heel width only.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "LOCKED 2026-04-30", "Status": "LOCKED",
                "Notes": "Replaces V1's overpromise that 'narrow heel volume' fixes shallow heel.",
                "ROMAN COMMENT": ""}),
        dict(ID="§3.P3.soft_mask", Trigger="Current shoe is soft (db_stiffness < 0.4) AND has perfect fit despite scan-shoe extreme mismatch",
             **{"Sentence template": "Your current shoes are soft enough to forgive some mismatch; stiffer recommendations may feel different at the same target geometry.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "LOCKED 2026-04-30", "Status": "LOCKED",
                "Notes": "Warns user that softness is hiding a fit risk visible only in stiffer shoes.",
                "ROMAN COMMENT": ""}),
        dict(ID="§3.P3.inconsistent", Trigger="Shoe-fit feedback inconsistent across multiple shoes (e.g., empty in one, tight in another)",
             **{"Sentence template": "Your shoe feedback was inconsistent across models. We weight the most consistent signal.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "LOCKED 2026-04-30", "Status": "LOCKED",
                "Notes": "Honest disclaimer when N>1 shoes give contradictory width or heel feedback.",
                "ROMAN COMMENT": ""}),
        dict(ID="§3.P3.tier_closing", Trigger="Always (closing line of P3)",
             **{"Sentence template": "The softer, stiffer, and budget tiers below offer alternative options around this target.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "LOCKED 2026-04-30", "Status": "LOCKED",
                "Notes": "Word 'adjacent' replaced with 'alternative' per Roman 2026-04-30.",
                "ROMAN COMMENT": ""}),
        # ============================================================
        # DERIVATION REFERENCE ROWS (substitution logic for P1 + P2)
        # ============================================================
        dict(ID="--- DERIVATIONS ---", Trigger="Substitution rules referenced by P1 + P2",
             **{"Sentence template": "(see following 5 rows)",
                "Concrete example": "",
                "Mentions shoe?": "n/a", "Origin": "REFERENCE", "Status": "LOCKED",
                "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="REF.discipline_phrase", Trigger="Builds the discipline_phrase substitution in §3.P2",
             **{"Sentence template": "indoor only -> '{discipline} indoors'; outdoor only -> '{discipline} outdoors on {rock}'; both equally -> '{discipline} both indoors and outdoors' (no rock since wizard skips that screen)",
                "Concrete example": "boulder + outdoor + sandstone -> 'bouldering outdoors on sandstone'\nboulder + both -> 'bouldering both indoors and outdoors'\ntrad_multipitch + indoor -> 'trad and multipitch indoors'",
                "Mentions shoe?": "No", "Origin": "LOCKED", "Status": "LOCKED",
                "Notes": "discipline labels: boulder -> 'bouldering'; sport -> 'sport climbing'; trad_multipitch -> 'trad and multipitch'.",
                "ROMAN COMMENT": ""}),
        dict(ID="REF.closure_pref", Trigger="Builds the closure_pref substitution in §3.P2 (discipline ∩ instep-allowed)",
             **{"Sentence template": "boulder × normal instep -> 'velcros or slippers'; boulder × low -> 'velcros'; boulder × high -> 'double-velcros'; sport × normal -> 'velcros or lace-ups'; sport × low -> 'double-velcros or lace-ups'; sport × high -> 'double-velcros or lace-ups'; trad/multipitch × any -> 'lace-ups'",
                "Concrete example": "Roman's case: boulder × low instep -> 'velcros' (slipper excluded). closure_caveat appended: '(avoiding slippers due to your low instep)'.",
                "Mentions shoe?": "No", "Origin": "LOCKED", "Status": "LOCKED",
                "Notes": "Anatomy exclusions: low instep excludes pure slippers. High instep excludes pure slippers AND single-strap velcros.",
                "ROMAN COMMENT": ""}),
        dict(ID="REF.stiffness_word", Trigger="Builds the stiffness_word substitution in §3.P2 (5 levels)",
             **{"Sentence template": "stiff_target < 0.25 -> 'very soft'; 0.25-0.40 -> 'soft'; 0.40-0.60 -> 'balanced stiffness'; 0.60-0.75 -> 'stiff'; >= 0.75 -> 'very stiff'",
                "Concrete example": "Roman's case: stiff_target = 0.30 -> 'soft'.",
                "Mentions shoe?": "No", "Origin": "LOCKED", "Status": "LOCKED",
                "Notes": "5 levels per Roman 2026-04-30. Note the 7-level vocabulary (super sensitive ... super supportive) still applies in shoe card P1 description (Roman didn't change that). 'Feel' axis remains dropped (V2 dead rule).",
                "ROMAN COMMENT": ""}),
        dict(ID="REF.downturn / asymmetry baseline", Trigger="Builds downturn + asymmetry baseline from aggressiveness",
             **{"Sentence template": "downturn: comfort -> 'flat profile'; balanced -> 'slight downturn'; moderate -> 'moderate downturn'; aggressive -> 'aggressive downturn'. asymmetry baseline: comfort -> 'no asymmetry'; balanced -> 'slight asymmetry'; moderate -> 'moderate asymmetry'; aggressive -> 'strong asymmetry'.",
                "Concrete example": "aggressive -> 'aggressive downturn' + 'strong asymmetry' (baseline before any foot-shape shift)",
                "Mentions shoe?": "No", "Origin": "LOCKED", "Status": "LOCKED",
                "Notes": "Baseline before foot-shape adjustment.", "ROMAN COMMENT": ""}),
        dict(ID="REF.asymmetry shift + tautology guard", Trigger="Builds final asymmetry target + asym_caveat from baseline + foot",
             **{"Sentence template": "Foot adjustment: Egyptian + no HVA -> +1 step (capped at 'strong'); HVA mild -> -1 step; HVA pronounced -> -2 steps. Tautology guard: caveat fires ONLY when target_asym_lbl != baseline_lbl. Caveat shifted-up wording: '(shifted higher to match your Egyptian toe shape)'. Caveat shifted-down: '(shifted lower to accommodate your {hva} hallux valgus)'.",
                "Concrete example": "Aggressive baseline='strong'. Egyptian + no HVA wants +1 but caps at strong -> still 'strong'. baseline_lbl == target_lbl -> NO caveat (Roman's case).\nMODERATE baseline + Egyptian + no HVA -> 'moderate'+1 = 'strong' -> caveat '(shifted higher ...)'.\nBALANCED baseline + HVA pronounced -> 'slight'-2 = 'no asymmetry' -> caveat '(shifted lower to accommodate your pronounced hallux valgus)'.",
                "Mentions shoe?": "No", "Origin": "LOCKED", "Status": "LOCKED",
                "Notes": "Tautology guard kept from earlier round.", "ROMAN COMMENT": ""}),
        # ============================================================
        # DROPPED V1 PARAGRAPHS (consolidated)
        # ============================================================
        dict(ID="--- DROPPED V1 ---", Trigger="V1 §3 paragraph variants superseded by the 3-paragraph design",
             **{"Sentence template": "(All V1 paragraph variants S3.1 through S3.14 are folded into the 3 templates above.)",
                "Concrete example": "",
                "Mentions shoe?": "n/a", "Origin": "V1 only", "Status": "DROPPED",
                "Notes": "Mapping summary: S3.1 + S3.2a/b/c -> §3.P1 (target stated as final, no explainer paragraphs).\nS3.3a/b/c/d/f + S3.4c/d -> folded into §3.P1 (target_fw / target_hv after adjustment shown directly, no clamped/unclamped explainer variants).\nS3.5a-e (cross-tier stiffness adjustments) -> folded into matrix_scorer_v2 logic; user-facing copy only mentions stiffness in §3.P2 once.\nS3.6 series (next_shoe_preference variants) -> already DROPPED (V2 wizard removed the input).\nS3.7 (forefoot paradox: squeezed + loose) -> kept as a per-shoe sentence in §2 S4.SQUEEZE_LOOSE; not duplicated in §3.\nS3.8 (toe form guidance) -> incorporated into §3.P1 toe_form target + §3.P2 closure recommendation.\nS3.9 closure paragraphs -> §3.P2 closure_pref logic.\nS3.10 tradeoffs -> covered in P3 conditional caveats (inconsistent feedback) and §1 T4 logic.\nS3.11 fit context -> implicit in §3.P1 (target IS the fit context).\nS3.12d (shallow heel transparency) -> §3.P3.shallow_heel (reworded).\nS3.13 (target shape) -> §3.P2.\nS3.14 / S3.14b (asym adjustment) -> folded into §3.P2 asym_caveat with tautology guard.",
                "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── Card P1 (UPDATED, drop P1.13) ─────────────────────────────
def build_card_p1(wb):
    ws = wb.create_sheet("Card P1 (description)")
    write_header(ws, "Shoe Card, Paragraph 1 (description)",
                 "Drop P1.13 (La Sportiva 'P3 system' boilerplate) per Roman 2026-04-28. Other rows unchanged.")

    rows = [
        dict(ID="P1.1", Trigger="width AND heel_volume defined",
             **{"Sentence template": "{Width} fit throughout{vol_suffix}",
                "Concrete example": "Narrow fit throughout, low forefoot volume",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "vol_suffix: ', low/high forefoot volume' or empty.", "ROMAN COMMENT": ""}),
        dict(ID="P1.2", Trigger="width defined, heel_volume absent",
             **{"Sentence template": "{Width} forefoot{vol_suffix}",
                "Concrete example": "Narrow forefoot",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P1.3", Trigger="closure == 'slipper'",
             **{"Sentence template": "Slipper construction",
                "Concrete example": "Narrow fit throughout, slipper construction.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Appended to fit_str.", "ROMAN COMMENT": ""}),
        dict(ID="P1.4", Trigger="closure == 'lace'",
             **{"Sentence template": "Lace closure",
                "Concrete example": "Medium fit throughout, lace closure.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P1.5", Trigger="toe_form list non-empty (V1, V2 STRIPS when matches user)",
             **{"Sentence template": "{Toe_forms_joined} toe box",
                "Concrete example": "Egyptian toe box.",
                "Mentions shoe?": "No", "Origin": "V1 (V2 strips matching)", "Status": "V2-DRAFT",
                "Notes": "V2 strips when shoe matches user's toe shape.", "ROMAN COMMENT": ""}),
        dict(ID="P1.6+7", Trigger="downturn defined (V1, V2 STRIPS when matches target)",
             **{"Sentence template": "{Downturn_label} OR {Downturn_label} with {Asymmetry_label}",
                "Concrete example": "Aggressively downturned with strong asymmetry.",
                "Mentions shoe?": "No", "Origin": "V1 (V2 strips matching)", "Status": "V2-DRAFT",
                "Notes": "V2 strips when sentence matches target_dt + target_asym pair.", "ROMAN COMMENT": ""}),
        dict(ID="P1.8", Trigger="no_edge == True",
             **{"Sentence template": "No-edge design wraps rubber smoothly around the sole for smearing but less precision on small edges",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Always kept (genuine shoe-specific characteristic).", "ROMAN COMMENT": ""}),
        dict(ID="P1.9", Trigger="rubber_type and/or rubber_thickness_mm exist",
             **{"Sentence template": "{thickness}mm {rubber_compound} rubber with a {midsole_coverage}, {midsole_stiffness} midsole, resulting in a {stiffness_word} shoe",
                "Concrete example": "3.5mm Vibram XS Grip rubber with a full, soft midsole, resulting in a sensitive shoe.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Roman locked sole format 2026-04-04.", "ROMAN COMMENT": ""}),
        dict(ID="P1.12", Trigger="stiffness computed exists",
             **{"Sentence template": "{stiffness_word} shoe",
                "Concrete example": "Balanced shoe.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "7 levels: super sensitive ... super supportive.", "ROMAN COMMENT": ""}),
        dict(ID="P1.13", Trigger="(was: description contains 'P3' La Sportiva)",
             **{"Sentence template": "(DROPPED per Roman 2026-04-28)",
                "Concrete example": "(N/A)",
                "Mentions shoe?": "-", "Origin": "V1 only", "Status": "DROPPED",
                "Notes": "Roman 2026-04-28: 'remove'. La Sportiva-specific boilerplate that doesn't add per-shoe value.",
                "ROMAN COMMENT": ""}),
        dict(ID="P1.14", Trigger="special_fit_notes field populated in DB",
             **{"Sentence template": "{special_fit_notes}",
                "Concrete example": "(arbitrary string)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Free-form per-shoe note.", "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── Card P2 (UPDATED) ──────────────────────────────────────────
def build_card_p2(wb):
    ws = wb.create_sheet("Card P2 (why selected)")
    write_header(ws, "Shoe Card, Paragraph 2 (why selected)",
                 "P2.3/P2.4 rewritten to reference toe SHAPE (Egyptian/Greek/Roman). P2.5/P2.6 softened. P2.7-8 dropped shallow+narrow heel conflation (we don't have shoe heel depth data, only width).")

    rows = [
        dict(ID="P2.1", Trigger="heel_empty score ≥3, <70% peers share",
             **{"Sentence template": "the tighter heel should fix the empty-heel feeling from your current shoe",
                "Concrete example": "Selected because the tighter heel should fix the empty-heel feeling from your current shoe.",
                "Mentions shoe?": "No (refers to 'current shoe')", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.2", Trigger="heel_tight score ≥3",
             **{"Sentence template": "the roomier heel relieves the tightness you felt in your current shoe",
                "Concrete example": "Selected because the roomier heel relieves the tightness you felt in your current shoe.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.3", Trigger="toes_squeezed score ≥3 (REVISED to reference toe SHAPE not size)",
             **{"Sentence template": "the {shoe_toe_form}-shaped toe box matches your {user_toe} foot, giving your toes more room than your current shoe",
                "Concrete example": "Selected because the Egyptian-shaped toe box matches your Egyptian foot, giving your toes more room than your current shoe.",
                "Mentions shoe?": "No", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "Roman 2026-04-28: 'this must refer to toe shape. We dont have info on toe box size, only if egyptian, roman, greek.' Old wording 'wider toe box' implied a width measurement we don't have.",
                "ROMAN COMMENT": ""}),
        dict(ID="P2.4", Trigger="toes_roomy score ≥3 (REVISED)",
             **{"Sentence template": "the {shoe_toe_form}-shaped toe box matches your {user_toe} foot, removing the dead space your current shoe leaves at your second toe",
                "Concrete example": "Selected because the Greek-shaped toe box matches your Greek foot, removing the dead space your current shoe leaves at your second toe.",
                "Mentions shoe?": "No", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "Same fix as P2.3. References toe shape match, not toe-box dimensions.",
                "ROMAN COMMENT": ""}),
        dict(ID="P2.5", Trigger="ff_tight score ≥3 (REVISED, softened)",
             **{"Sentence template": "the wider forefoot may relieve the tightness you felt in your current shoe",
                "Concrete example": "Selected because the wider forefoot may relieve the tightness you felt in your current shoe.",
                "Mentions shoe?": "No", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "Roman 2026-04-28: 'sounds too certain. Reformulate'. Changed 'relieves' to 'may relieve'.",
                "ROMAN COMMENT": ""}),
        dict(ID="P2.6", Trigger="ff_loose score ≥3 (REVISED, softened)",
             **{"Sentence template": "the snugger forefoot may fix the loose feeling you experienced",
                "Concrete example": "Selected because the snugger forefoot may fix the loose feeling you experienced.",
                "Mentions shoe?": "No", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "Same softening as P2.5.",
                "ROMAN COMMENT": ""}),
        dict(ID="P2.7", Trigger="user heel_width=='narrow' AND shoe.db_heel_volume=='narrow' (REVISED)",
             **{"Sentence template": "the narrow heel cup matches your narrow heel and should fix the empty-heel feeling from your current shoes",
                "Concrete example": "Selected because the narrow heel cup matches your narrow heel and should fix the empty-heel feeling from your current shoes.",
                "Mentions shoe?": "No", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "Roman 2026-04-28: 'WTF we mix shallow and narrow hell??? Before you wrote that we don't have info on heel depth for shoes'. Fix: removed shallow heel claim entirely. References ONLY heel WIDTH match (which we have data for).",
                "ROMAN COMMENT": ""}),
        dict(ID="P2.8", Trigger="user heel_width=='narrow' AND shoe.db_heel_volume=='narrow', no other heel reason (REVISED)",
             **{"Sentence template": "the narrow heel cup matches your narrow heel",
                "Concrete example": "Selected because the narrow heel cup matches your narrow heel.",
                "Mentions shoe?": "No", "Origin": "REVISED", "Status": "REVISED",
                "Notes": "Standalone narrow-heel match. No shallow conflation.",
                "ROMAN COMMENT": ""}),
        dict(ID="P2.9", Trigger="instep_downturn score ≥3",
             **{"Sentence template": "the downturn relieves pressure on your high instep",
                "Concrete example": "Selected because the downturn relieves pressure on your high instep.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.10", Trigger="stiffness differs from peer avg by ±0.10+",
             **{"Sentence template": "noticeably {stiffer|softer} than most other picks, for more {support|sensitivity}",
                "Concrete example": "Selected because noticeably stiffer than most other picks, for more support.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Tier standout, fires often.", "ROMAN COMMENT": ""}),
        dict(ID="P2.12", Trigger="downturn rank differs from peer avg by >0.8",
             **{"Sentence template": "the {dt_word} downturn is {more|less} aggressive than most picks, {built for steeper terrain|trading power for comfort}",
                "Concrete example": "Selected because the slight downturn is less aggressive than most picks, trading power for comfort.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.14", Trigger="feel differs from peer avg by >1.2 ranks",
             **{"Sentence template": "the {shoe_feel} feel is {softer|firmer} than most other picks",
                "Concrete example": "Selected because the sensitive feel is softer than most other picks.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.16", Trigger="width stands out vs peers (±0.5 rank)",
             **{"Sentence template": "the {wider|narrower} forefoot {gives your toes more room|wraps tighter} than most other picks",
                "Concrete example": "Selected because the wider forefoot gives your toes more room than most other picks.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.17-18", Trigger="no_edge == True (only / shared)",
             **{"Sentence template": "the only edgeless shoe in this set, wrapping around footholds for extra contact (only) OR edgeless design wraps around footholds for extra contact (shared)",
                "Concrete example": "Selected because edgeless design wraps around footholds for extra contact.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.19", Trigger="closure differs from peer majority",
             **{"Sentence template": "the {lace-up closure|slipper design|velcro closure} {allows custom fit|gives low-profile sensitive fit|allows quick on/off}",
                "Concrete example": "Selected because the lace-up closure allows a more custom fit.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.23", Trigger="No specific reason fires (V1 fallback, DROPPED in V2)",
             **{"Sentence template": "Good overall fit for your foot shape and climbing style.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V1 only", "Status": "DROPPED",
                "Notes": "DROPPED in V2 (returns None). Showed 4/12 times in Roman's review.", "ROMAN COMMENT": ""}),
        dict(ID="P2.24", Trigger="not_in_stock == True",
             **{"Sentence template": "Note: this shoe is not currently available online. Check local shops or wait for restocks.",
                "Concrete example": "(appended to P2)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P2.25", Trigger="tier == 'budget' AND best_price exists",
             **{"Sentence template": "At EUR {price:.0f}, this is a strong value pick.",
                "Concrete example": "At EUR 124, this is a strong value pick.",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "V1-LIVE",
                "Notes": "Slightly redundant with size pill (which already shows EUR). Worth reviewing whether to drop in V2.",
                "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── Card P3 (audit pass, no major changes flagged) ────────────
def build_card_p3(wb):
    ws = wb.create_sheet("Card P3 (tradeoffs)")
    write_header(ws, "Shoe Card, Paragraph 3 (tradeoffs)",
                 "Audit pass after Roman's review, no specific changes flagged. V2 already drops the 'No notable tradeoffs...' fallback.")

    rows = [
        dict(ID="P3.2", Trigger="forefoot_width axis < 0 (V2)",
             **{"Sentence template": "the {shoe_width} forefoot runs {wider|narrower} than your {target_label} target",
                "Concrete example": "Tradeoff: The medium forefoot runs wider than your narrow target.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P3.4", Trigger="heel_volume axis < 0 (V2)",
             **{"Sentence template": "the {shoe_hv} heel volume is {roomier|tighter} than your {target_label} target",
                "Concrete example": "Tradeoff: The medium heel volume is roomier than your narrow target.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P3.7", Trigger="toe_form axis < 0 (V2), user_toe not in shoe forms",
             **{"Sentence template": "the toe box is shaped for {shoe_forms} feet, not {user_toe} (Greek case) OR built for {shoe_forms} feet, opposite to your {user_toe} foot (Egyptian/Roman)",
                "Concrete example": "Tradeoff: The toe box is shaped for Egyptian feet, not Greek.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT",
                "Notes": "V2: Greek mismatch -5×conf, Egyptian/Roman opposites -10×conf.", "ROMAN COMMENT": ""}),
        dict(ID="P3.15", Trigger="downturn axis < 0 (V2)",
             **{"Sentence template": "the {dt_short} downturn is {less|more} aggressive than the {target_label} downturn we target for your selection",
                "Concrete example": "Tradeoff: The slight downturn is less aggressive than the aggressive downturn we target for your selection.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P3.18", Trigger="asymmetry axis < 0 (V2)",
             **{"Sentence template": "the {shoe_asym} asymmetry is {more|less} {aggressive|pronounced} than the {target_label} asymmetry we target",
                "Concrete example": "Tradeoff: The slight asymmetry is less pronounced than the strong asymmetry we target.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT",
                "Notes": "Target_asym is foot-shape adjusted.", "ROMAN COMMENT": ""}),
        dict(ID="P3.20", Trigger="stiffness axis < 0 (V2), tier-aware",
             **{"Sentence template": "the {stiff_word} sole is {softer|stiffer} than what you currently climb in (with anchor) OR sits outside the comfortable range for your selection (no anchor)",
                "Concrete example": "Tradeoff: The supportive sole is stiffer than what you currently climb in.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT",
                "Notes": "Tier-aware: softer tier doesn't surface 'stiffer than current'.", "ROMAN COMMENT": ""}),
        dict(ID="P3.23", Trigger="instep_extreme axis < 0, shoe_closure == 'slipper', user_instep not normal",
             **{"Sentence template": "the slipper closure leaves no adjustability for your {user_instep_clean} instep",
                "Concrete example": "Tradeoff: The slipper closure leaves no adjustability for your low instep.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P3.39", Trigger="discipline_overlap axis < 0 (V2)",
             **{"Sentence template": "the shoe is not built for {discipline} climbing",
                "Concrete example": "Tradeoff: The shoe is not built for trad climbing.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT",
                "Notes": "V2: -100 score, overrides all. Filters out unsuitable picks.", "ROMAN COMMENT": ""}),
        dict(ID="P3.40", Trigger="closure axis < 0 (V2)",
             **{"Sentence template": "the {shoe_closure} closure is not ideal for {discipline} climbing",
                "Concrete example": "Tradeoff: The slipper closure is not ideal for trad/multipitch climbing.",
                "Mentions shoe?": "No", "Origin": "V2", "Status": "V2-DRAFT", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="P3.42", Trigger="No tradeoff issues at all (V1 fallback, DROPPED in V2)",
             **{"Sentence template": "No notable tradeoffs for your foot shape.",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "V1 only", "Status": "DROPPED",
                "Notes": "DROPPED in V2 (returns None, card omits P3 entirely). Showed 10/12 times in Roman's review.",
                "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


# ─── Cross-cutting (unchanged) ──────────────────────────────────
def build_cross(wb):
    ws = wb.create_sheet("Cross-cutting")
    write_header(ws, "Cross-cutting, sizing, tier headers, anchors",
                 "Shared elements that appear across all results pages. No changes flagged in Roman's review.")

    rows = [
        dict(ID="X.SIZE", Trigger="Per-shoe (every recommendation card)",
             **{"Sentence template": "EU {recommended_size_eu}",
                "Concrete example": "EU 44 (size pill in shoe card header)",
                "Mentions shoe?": "Implicit", "Origin": "V1 + V2", "Status": "LOCKED",
                "Notes": "MUST snap to half-EU via round(x*2)/2.", "ROMAN COMMENT": ""}),
        dict(ID="X.PRICE", Trigger="Budget tier shoes only",
             **{"Sentence template": "{price:.0f} EUR @ {retailer}",
                "Concrete example": "124 EUR @ TheOutlet (in size pill)",
                "Mentions shoe?": "No", "Origin": "V1 + V2", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="X.TIER.baseline", Trigger="Tier header (always shown if tier has picks)",
             **{"Sentence template": "Your Best Match : Similar feel and use case to your current shoes",
                "Concrete example": "(verbatim header)",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED",
                "Notes": "From CATEGORY_META.", "ROMAN COMMENT": ""}),
        dict(ID="X.TIER.softer", Trigger="Softer tier header",
             **{"Sentence template": "Softer Shoes : For more sensitivity, recommended for indoors and bouldering",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="X.TIER.stiffer", Trigger="Stiffer tier header",
             **{"Sentence template": "Stiffer Shoes : For more support, recommended for outdoors and sport/trad climbing",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="X.TIER.budget", Trigger="Budget tier header",
             **{"Sentence template": "Best Value : Affordable picks at your recommended size",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="X.SEEMORE", Trigger="Per tier (after all picks shown)",
             **{"Sentence template": "See more {top matches | softer picks | stiffer picks | value picks} →",
                "Concrete example": "See more top matches →",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED",
                "Notes": "Links to /scan/{scan_id}/browse?tier={cat}.", "ROMAN COMMENT": ""}),
        dict(ID="X.SHOE.CTA", Trigger="Per shoe card (always)",
             **{"Sentence template": "Check details and availability",
                "Concrete example": "(button at bottom of every shoe card)",
                "Mentions shoe?": "Implicit", "Origin": "Live ScanResult.jsx", "Status": "LOCKED", "Notes": "", "ROMAN COMMENT": ""}),
        dict(ID="X.HEADER", Trigger="Page header (always)",
             **{"Sentence template": "climbing-gear.com / Your Foot Profile / Scan analysis: sole & side view",
                "Concrete example": "(verbatim)",
                "Mentions shoe?": "No", "Origin": "Live ScanResult.jsx", "Status": "LOCKED",
                "Notes": "Em dash removed from subtitle.", "ROMAN COMMENT": ""}),
    ]
    start = 5
    for i, r in enumerate(rows):
        add_row(ws, start + i, r)
    auto_row_height(ws, start, start + len(rows) - 1)


def main():
    wb = Workbook()
    build_cover(wb)
    build_section1(wb)
    build_section2(wb)
    build_section3(wb)
    build_card_p1(wb)
    build_card_p2(wb)
    build_card_p3(wb)
    build_cross(wb)
    wb.save(OUT)
    print(f"# wrote {OUT} ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
